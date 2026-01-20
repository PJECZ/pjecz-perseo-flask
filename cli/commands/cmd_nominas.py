"""
CLI Nominas
"""

import csv
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import click
import xlrd
from dotenv import load_dotenv
from openpyxl import load_workbook

from pjecz_perseo_flask.blueprints.centros_trabajos.models import CentroTrabajo
from pjecz_perseo_flask.blueprints.conceptos.models import Concepto
from pjecz_perseo_flask.blueprints.nominas.generators.aguinaldos import crear_aguinaldos
from pjecz_perseo_flask.blueprints.nominas.generators.dispersiones_pensionados import crear_dispersiones_pensionados
from pjecz_perseo_flask.blueprints.nominas.generators.monederos import crear_monederos
from pjecz_perseo_flask.blueprints.nominas.generators.nominas import crear_nominas
from pjecz_perseo_flask.blueprints.nominas.generators.pensionados import crear_pensionados
from pjecz_perseo_flask.blueprints.nominas.generators.primas_vacacionales import crear_primas_vacacionales
from pjecz_perseo_flask.blueprints.nominas.generators.timbrados import crear_timbrados
from pjecz_perseo_flask.blueprints.nominas.models import Nomina
from pjecz_perseo_flask.blueprints.percepciones_deducciones.models import PercepcionDeduccion
from pjecz_perseo_flask.blueprints.personas.models import Persona
from pjecz_perseo_flask.blueprints.plazas.models import Plaza
from pjecz_perseo_flask.blueprints.puestos.models import Puesto
from pjecz_perseo_flask.blueprints.quincenas.models import Quincena
from pjecz_perseo_flask.blueprints.quincenas_productos.models import QuincenaProducto
from pjecz_perseo_flask.blueprints.tabuladores.models import Tabulador
from pjecz_perseo_flask.blueprints.timbrados.models import Timbrado
from pjecz_perseo_flask.config.extensions import database
from pjecz_perseo_flask.lib.fechas import crear_clave_quincena, quincena_to_fecha, quinquenio_count
from pjecz_perseo_flask.lib.safe_string import QUINCENA_REGEXP, safe_clave, safe_quincena, safe_rfc, safe_string
from pjecz_perseo_flask.main import app

load_dotenv()

EXPLOTACION_BASE_DIR = os.getenv("EXPLOTACION_BASE_DIR", "")
EXTRAORDINARIOS_BASE_DIR = os.getenv("EXTRAORDINARIOS_BASE_DIR", "")
PENSIONES_ALIMENTICIAS_BASE_DIR = os.getenv("PENSIONES_ALIMENTICIAS_BASE_DIR", "")

AGUINALDOS_FILENAME_XLS = "Aguinaldos.XLS"
APOYOS_FILENAME_XLS = "Apoyos.XLS"
APOYOS_DIA_DE_LA_MADRE_FILENAME_XLS = "Apoyos_dia_de_la_madre.XLS"
BONOS_FILENAME_XLS = "Bonos.XLS"
COMPANIA_NOMBRE = "PODER JUDICIAL DEL ESTADO DE COAHUILA DE ZARAGOZA"
COMPANIA_RFC = "PJE901211TI9"
COMPANIA_CP = "25000"
NOMINAS_FILENAME_XLS = "NominaFmt2.XLS"
PATRON_RFC = "PJE901211TI9"
PRIMAS_FILENAME_XLS = "PrimaVacacional.XLS"
RETROACTIVOS_AGUINALDOS_FILENAME_XLS = "RetroactivosAguinaldos.XLS"
SALARIOS_RL_FILENAME_XLS = "SalariosRL.XLS"
SERICA_FILENAME_XLSX = "SERICA.xlsx"

# Inicializar el contexto de la aplicación Flask
app.app_context().push()


@click.group()
def cli():
    """Nominas"""


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("archivo_csv", type=str)
@click.option("--tipo", type=str, default="SALARIO")
@click.option("--vaciar_primero", is_flag=True, default=False, help="Primero pone textos vacios.")
@click.option("--probar", is_flag=True, help="Solo probar sin cambiar la base de datos.")
def actualizar_numeros_cheque(quincena_clave: str, archivo_csv: str, tipo: str, vaciar_primero: bool, probar: bool = False):
    """Actualizar los numeros de cheque a partir de un CSV con 'RFC' y 'NO DE CHEQUE'"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena, se termina
    if quincena is None:
        click.echo(f"ERROR: Quincena {quincena_clave} no existe.")
        sys.exit(1)

    # Validar archivo CSV
    ruta = Path(archivo_csv)
    if not ruta.exists():
        click.echo(f"ERROR: {ruta.name} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {ruta.name} no es un archivo.")
        sys.exit(1)

    # Validar tipo
    tipo = safe_string(tipo)
    if tipo not in Nomina.TIPOS:
        click.echo("ERROR: Tipo inválido.")
        sys.exit(1)

    # Si se pide vaciar_primero todos los numeros de cheque
    actualizados_contador = 0
    if vaciar_primero is True:
        if probar is True:
            click.echo("[PROBAR] ", nl=False)
        click.echo(f"Poniendo en texto vacio el numero de cheque de TODAS las nominas {quincena_clave} y {tipo}: ", nl=False)
        nominas = (
            Nomina.query.join(Quincena)
            .filter(Quincena.clave == quincena_clave)
            .filter(Nomina.tipo == tipo)
            .filter(Nomina.estatus == "A")
            .all()
        )
        for nomina in nominas:
            if nomina.num_cheque != "":
                nomina.num_cheque = ""
                actualizados_contador += 1
                if probar is False:
                    nomina.save()
                click.echo(click.style("u", fg="cyan"), nl=False)
        click.echo()
        if probar is True:
            click.echo("[PROBAR] ", nl=False)
        click.echo(click.style(f"  Se pusieron en texto vacio {actualizados_contador} numeros de cheque", fg="green"))

    # Leer el archivo CSV
    actualizados_contador = 0
    no_encontrados_contador = 0
    if probar is True:
        click.echo("[PROBAR] ", nl=False)
    click.echo(f"Actualizando los numeros de cheque las nominas {quincena_clave} y {tipo}: ", nl=False)
    with open(ruta, encoding="utf8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            rfc = safe_rfc(row["RFC"])
            num_cheque = row["NO DE CHEQUE"]
            # Consultar la Nomina
            nomina = (
                Nomina.query.join(Persona)
                .filter(Persona.rfc == rfc)
                .filter(Nomina.quincena_id == quincena.id)
                .filter(Nomina.tipo == tipo)
                .first()
            )
            # Si existe la Nomina, se actualiza el numero de cheque
            if nomina is not None:
                if nomina.num_cheque != num_cheque:
                    nomina.num_cheque = num_cheque
                    actualizados_contador += 1
                    if probar is False:
                        nomina.save()
                    click.echo(click.style(".", fg="cyan"), nl=False)
            else:
                no_encontrados_contador += 1
                click.echo(click.style("X", fg="yellow"), nl=False)
        click.echo()

    # Si no_encontrados_contador es mayor a cero, mostrar
    if no_encontrados_contador > 0:
        if probar is True:
            click.echo("[PROBAR] ", nl=False)
        click.echo(click.style(f"  No se encontraron {no_encontrados_contador} RFCs", fg="yellow"))

    # Mensaje termino
    if probar is True:
        click.echo("[PROBAR] ", nl=False)
    click.echo(click.style(f"  Se actualizaron {actualizados_contador} numeros de cheque", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def actualizar_timbrados(quincena_clave: str, probar: bool = False):
    """Actualizar los timbrados_ids (cambiar a cero si esta eliminado o no existe) de las nominas de una quincena"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena, se termina
    if quincena is None:
        click.echo(f"ERROR: Quincena {quincena_clave} no existe.")
        sys.exit(1)

    # Inicializar contadores
    revisados_contador = 0
    actualizados_contador = 0
    puestos_en_cero_contador = 0

    # Bucle por cada nomina de la quincena
    click.echo(f"Actualizando los timbrados_ids de la quincena {quincena_clave}: ", nl=False)
    for nomina in quincena.nominas:
        # Si el estatus NO es 'A', se omite
        if nomina.estatus != "A":
            continue

        # Si el tipo es 'DESPENSA', se omite
        if nomina.tipo == "DESPENSA":
            continue

        # Incrementar revisados_contador
        revisados_contador += 1

        # De forma inicial, se define timbrado_id en None
        timbrado_id = None

        # Consultar el ultimo timbrado de la Nomina con estatus 'A' y ordenados por id descendente
        timbrado_ultimo = (
            Timbrado.query.filter_by(nomina_id=nomina.id).filter_by(estatus="A").order_by(Timbrado.id.desc()).first()
        )

        # Si existe el timbrado_ultimo, se toma su id
        if timbrado_ultimo is not None:
            timbrado_id = timbrado_ultimo.id

        # Si el timbrado_id es diferente al timbrado_id de la nomina, se actualiza
        if nomina.timbrado_id != timbrado_id:
            nomina.timbrado_id = timbrado_id
            if probar is False:
                nomina.save()
            actualizados_contador += 1
            # Si NO hay timbrado entonces timbrado_id es cero, se incrementa el contador
            if timbrado_id is None:
                puestos_en_cero_contador += 1
                click.echo(click.style("0", fg="yellow"), nl=False)
            else:
                click.echo(click.style("u", fg="green"), nl=False)
        else:
            click.echo(click.style(".", fg="cyan"), nl=False)

    # Poner avance de linea en la terminal
    click.echo("")

    # Si revisados_contador es cero, mostrar mensaje de error y terminar
    if revisados_contador == 0:
        click.echo(click.style("  No hubo registros de Nominas que revisar", fg="red"))
        sys.exit(1)

    # Si hubo nominas actualizadas, mostrar contador
    if actualizados_contador > 0:
        click.echo(click.style(f"  Se ACTUALIZARON {actualizados_contador} Nominas", fg="green"))

    # Si hubo puestos_en_cero_contador en cero, mostrar contador
    if puestos_en_cero_contador > 0:
        click.echo(click.style(f"  De la cuales {puestos_en_cero_contador} Nominas cambiaron timbrado_id a cero", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Se revisaron {revisados_contador} Nominas", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar nominas"""

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Definir la fecha_final en base a la clave de la quincena
    try:
        fecha_final = quincena_to_fecha(quincena_clave, dame_ultimo_dia=True)
    except ValueError:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, NOMINAS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Definir el puesto generico al que se van a relacionar las personas que no tengan su puesto
    puesto_generico = Puesto.query.filter_by(clave="ND").first()
    if puesto_generico is None:
        click.echo("ERROR: Falta el puesto con clave ND.")
        sys.exit(1)

    # Definir el tabulador generico al que se van a relacionar los puestos que no tengan su tabulador
    tabulador_generico = Tabulador.query.filter_by(puesto_id=puesto_generico.id).first()
    if tabulador_generico is None:
        click.echo("ERROR: Falta el tabulador del puesto con clave ND.")
        sys.exit(1)

    # Inicializar los listados con las anomalias
    personas_actualizadas_del_tabulador = []
    personas_actualizadas_del_modelo = []
    personas_actualizadas_del_num_empleado = []
    personas_sin_puestos = []
    personas_sin_tabulador = []

    # Iniciar contadores
    contador = 0
    centros_trabajos_insertados_contador = 0
    personas_actualizadas_contador = 0
    personas_insertadas_contador = 0
    plazas_insertadas_contador = 0

    # Bucle por cada fila
    click.echo(f"Alimentar Nominas a la quincena {quincena.clave}: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        centro_trabajo_clave = hoja.cell_value(fila, 1)
        plaza_clave = hoja.cell_value(fila, 8)
        percepcion = int(hoja.cell_value(fila, 12)) / 100.0
        deduccion = int(hoja.cell_value(fila, 13)) / 100.0
        impte = int(hoja.cell_value(fila, 14)) / 100.0
        desde_s = str(int(hoja.cell_value(fila, 16)))
        hasta_s = str(int(hoja.cell_value(fila, 17)))

        # Tomar las columnas con datos de la Persona
        rfc = hoja.cell_value(fila, 2)
        modelo = int(hoja.cell_value(fila, 236))
        nombre_completo = hoja.cell_value(fila, 3)
        num_empleado = int(hoja.cell_value(fila, 240))

        # Tomar las columnas necesarias para el timbrado
        puesto_clave = safe_clave(hoja.cell_value(fila, 20))
        nivel = int(hoja.cell_value(fila, 9))
        quincena_ingreso = str(int(hoja.cell_value(fila, 19)))

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Consultar el Centro de Trabajo, si no existe se agrega
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()
        if centro_trabajo is None:
            centro_trabajo = CentroTrabajo(clave=centro_trabajo_clave, descripcion="ND")
            sesion.add(centro_trabajo)
            sesion.commit()
            centros_trabajos_insertados_contador += 1

        # Consultar la Plaza, si no existe se agrega
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()
        if plaza is None:
            plaza = Plaza(clave=plaza_clave, descripcion="ND")
            sesion.add(plaza)
            sesion.commit()
            plazas_insertadas_contador += 1

        # Si el modelo es 2, entonces en SINDICALIZADO, se toman 4 caracteres del puesto y se busca quinquenios
        quinquenios = None
        if modelo == 2:
            puesto_clave = puesto_clave[:4]

            # Inicializar la bandera para saltar la fila si el concepto es PME
            es_concepto_pme = False

            # Bucle entre las columnas de los conceptos para encontrar PQ1, PQ2, PQ3, PQ4, PQ5, PQ6
            col_num = 26
            while True:
                # Tomar el p_o_d
                p_o_d = safe_string(hoja.cell_value(fila, col_num))
                # Tomar el conc
                conc = safe_string(hoja.cell_value(fila, col_num + 1))
                # Si 'P' o 'D' es un texto vacio, se rompe el ciclo
                if p_o_d == "":
                    break
                # Si NO es P, se salta
                if p_o_d != "P":
                    col_num += 6
                    continue
                # Si conc es ME es monedero, se rompe el ciclo porque aqui no hay quinquenios
                if conc == "ME":
                    es_concepto_pme = True
                    break
                # Si el concepto no es PQ1, PQ2, PQ3, PQ4, PQ5, PQ6, se salta
                if conc not in ["Q1", "Q2", "Q3", "Q4", "Q5", "Q6"]:
                    col_num += 6
                    continue
                # Tomar el tercer caracter del concepto y convertirlo a entero porque es la cantidad de quinquenios
                quinquenios = int(conc[1])
                break

            # Si es PME, entonces en esta fila NO esta el quinquenio, se mantiene en None
            if es_concepto_pme:
                quinquenios = None

        else:
            # Entonces NO es SINDICALIZADO, se define quinquenios en cero
            quinquenios = 0

        # Consultar el Puesto, si no existe se agrega a personas_sin_puestos y se le asigna el puesto_generico
        puesto = Puesto.query.filter_by(clave=puesto_clave).first()
        if puesto is None:
            personas_sin_puestos.append(rfc)
            puesto = puesto_generico

        # Consultar la Persona
        persona = Persona.query.filter_by(rfc=rfc).first()

        # Si NO existe la Persona, se agrega
        if persona is None:
            # Separar nombre_completo, en apellido_primero, apellido_segundo y nombres
            separado = safe_string(nombre_completo, save_enie=True).split(" ")
            apellido_primero = separado[0]
            apellido_segundo = separado[1]
            nombres = " ".join(separado[2:])

            # Si el modelo es 2 y quinquenios es None, entonces es SINDICALIZADO y se calculan los quinquenios
            if modelo == 2 and quinquenios is None:
                # Calcular la cantidad de quinquenios
                fecha_ingreso = quincena_to_fecha(quincena_ingreso, dame_ultimo_dia=False)
                quinquenios = quinquenio_count(fecha_ingreso, fecha_final)

            # Consultar el tabulador que coincida con puesto_clave, modelo, nivel y quinquenios
            tabulador = (
                Tabulador.query.filter_by(puesto_id=puesto.id)
                .filter_by(modelo=modelo)
                .filter_by(nivel=nivel)
                .filter_by(quinquenio=quinquenios)
                .first()
            )

            # Si no existe el tabulador, se agrega a personas_sin_tabulador y se le asigna tabulador_generico
            if tabulador is None:
                personas_sin_tabulador.append(rfc)
                tabulador = tabulador_generico

            # Insertar a la Persona
            persona = Persona(
                tabulador_id=tabulador.id,
                rfc=rfc,
                nombres=nombres,
                apellido_primero=apellido_primero,
                apellido_segundo=apellido_segundo,
                modelo=modelo,
                num_empleado=num_empleado,
            )
            sesion.add(persona)
            sesion.commit()
            personas_insertadas_contador += 1

        # De lo contrario, se revisa si cambia la Persona de tabulador, modelo o num_empleado
        else:
            # Inicializar hay_cambios
            hay_cambios = False

            # Si la fila es concepto PME NO va tener los quinquenios, entonces se define con la Persona
            if quinquenios is None:
                quinquenios = persona.tabulador.quinquenio

            # Consultar el tabulador que coincida con puesto_clave, modelo, nivel y quinquenios
            tabulador = (
                Tabulador.query.filter_by(puesto_id=puesto.id)
                .filter_by(modelo=modelo)
                .filter_by(nivel=nivel)
                .filter_by(quinquenio=quinquenios)
                .first()
            )

            # Si NO existe el tabulador, se agrega a personas_sin_tabulador y se le asigna tabulador_generico
            if tabulador is None:
                personas_sin_tabulador.append(rfc)
                tabulador = tabulador_generico

            # Revisar si hay que actualizar el tabulador a la Persona
            if persona.tabulador_id != tabulador.id:
                personas_actualizadas_del_tabulador.append(
                    f"{rfc} {persona.nombre_completo}: Tabulador: {persona.tabulador_id} -> {tabulador.id}"
                )
                persona.tabulador_id = tabulador.id
                hay_cambios = True

            # Revisar si hay que actualizar el modelo a la Persona
            if persona.modelo != modelo:
                personas_actualizadas_del_modelo.append(
                    f"{rfc} {persona.nombre_completo}: Modelo: {persona.modelo} -> {modelo}"
                )
                persona.modelo = modelo
                hay_cambios = True

            # Revisar si hay que actualizar el numero de empleado a la Persona
            if persona.num_empleado != num_empleado:
                personas_actualizadas_del_num_empleado.append(
                    f"{rfc} {persona.nombre_completo}: Num. Emp. {persona.num_empleado} -> {num_empleado}"
                )
                persona.num_empleado = num_empleado
                hay_cambios = True

            # Si hay cambios, guardar la Persona
            if hay_cambios:
                if probar is False:
                    persona.save()
                personas_actualizadas_contador += 1

        # Bucle entre P-D para determinar el tipo entre SALARIO y DESPENSA
        nomina_tipo = None
        col_num = 26
        while True:
            # Tomar el tipo y el conc para armar la clave del concepto
            tipo = safe_string(hoja.cell_value(fila, col_num))
            conc = safe_string(hoja.cell_value(fila, col_num + 1))
            concepto_clave = f"{tipo}{conc}"

            # Si el tipo es un texto vacio, se rompe el ciclo
            if tipo == "":
                break

            # Si el concepto_clave es PME, entonces es DESPENSA y se termina este ciclo
            if concepto_clave == "PME":
                nomina_tipo = Nomina.TIPOS["DESPENSA"]
                break

            # Incrementar col_num en SEIS
            col_num += 6

            # Romper el ciclo cuando se llega a la columna
            if col_num > 236:
                break

        # Si no se encontro el tipo, entonces es SALARIO
        if nomina_tipo is None:
            nomina_tipo = Nomina.TIPOS["SALARIO"]

        # Alimentar nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo=nomina_tipo,
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)

        # Incrementar contador
        contador += 1

        # Mostrar el avance con el modelo
        click.echo(click.style(modelo, fg="cyan"), nl=False)

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    if probar is False:
        sesion.commit()
        sesion.close()

    # Si hubo centros_trabajos_insertados, mostrar contador
    if centros_trabajos_insertados_contador > 0:
        click.echo(click.style(f"  Se insertaron {centros_trabajos_insertados_contador} Centros de Trabajo", fg="green"))

    # Si hubo personas actualizadas, mostrar contador
    if personas_actualizadas_contador > 0:
        click.echo(click.style(f"  Se actualizaron {personas_actualizadas_contador} Personas", fg="green"))
        for item in personas_actualizadas_del_tabulador:
            click.echo(click.style(f"  {item}", fg="yellow"))
        for item in personas_actualizadas_del_modelo:
            click.echo(click.style(f"  {item}", fg="yellow"))
        for item in personas_actualizadas_del_num_empleado:
            click.echo(click.style(f"  {item}", fg="yellow"))

    # Si hubo personas insertadas, mostrar contador
    if personas_insertadas_contador > 0:
        click.echo(click.style(f"  Se insertaron {personas_insertadas_contador} Personas", fg="green"))

    # Si hubo plazas insertadas, mostrar contador
    if plazas_insertadas_contador > 0:
        click.echo(click.style(f"  Se insertaron {plazas_insertadas_contador} Plazas", fg="green"))

    # Si hubo personas_sin_puestos, mostrarlas en pantalla
    if len(personas_sin_puestos) > 0:
        click.echo(click.style(f"  Hubo {len(personas_sin_puestos)} Personas sin puestos.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_sin_puestos)}", fg="yellow"))

    # Si hubo personas_sin_tabulador, mostrarlas en pantalla
    if len(personas_sin_tabulador) > 0:
        click.echo(click.style(f"  Hubo {len(personas_sin_tabulador)} Personas sin tabulador.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_sin_tabulador)}", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Alimentar Nominas: {contador} insertadas.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_aguinaldos(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar aguinaldos"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Definir la fecha_final en base a la clave de la quincena
    # Va a tomar el año de la quincena, el mes 12 y el dia 31
    try:
        fecha = quincena_to_fecha(quincena_clave, dame_ultimo_dia=True)
        fecha_final = datetime(fecha.year, 12, 31)
    except ValueError:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesión con la base de datos para que la alimentación sea rápida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotación
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, AGUINALDOS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    personas_inexistentes = []
    plazas_insertadas_contador = 0

    # Bucle por cada fila
    click.echo(f"Alimentando Aguinaldos a la quincena {quincena.clave}: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        centro_trabajo_clave = hoja.cell_value(fila, 1)
        rfc = hoja.cell_value(fila, 2)
        # nombre_completo = hoja.cell_value(fila, 3)
        plaza_clave = hoja.cell_value(fila, 8)
        percepcion = int(hoja.cell_value(fila, 12)) / 100.0
        deduccion = int(hoja.cell_value(fila, 13)) / 100.0
        impte = int(hoja.cell_value(fila, 14)) / 100.0
        desde_s = str(int(hoja.cell_value(fila, 16)))
        hasta_s = str(int(hoja.cell_value(fila, 17)))
        # modelo = int(hoja.cell_value(fila, 236))
        # num_empleado = int(hoja.cell_value(fila, 240))

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Consultar la persona, si no existe, se agrega a la lista de personas_inexistentes y se salta
        persona = Persona.query.filter_by(rfc=rfc).first()
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo, si no existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Revisar si la Plaza existe, de lo contrario insertarla
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()
        if plaza is None:
            plaza = Plaza(clave=plaza_clave, descripcion="ND")
            sesion.add(plaza)
            plazas_insertadas_contador += 1

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo="AGUINALDO",
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)

        # Incrementar contador
        contador += 1
        if contador % 100 == 0:
            click.echo(click.style(".", fg="cyan"), nl=False)

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Actualizar la quincena para poner en verdadero el campo tiene_aguinaldos
    quincena.tiene_aguinaldos = True
    quincena.save()

    # Si hubo centros trabajos inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas insertadas, mostrar contador
    if plazas_insertadas_contador > 0:
        click.echo(click.style(f"  Se insertaron {plazas_insertadas_contador} Plazas", fg="green"))

    # Si hubo personas inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Alimentar Aguinaldos: {contador} insertadas en la quincena {quincena_clave}.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_apoyos_anuales(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar apoyos anuales"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Definir la fecha_final en base a la clave de la quincena
    try:
        fecha_final = quincena_to_fecha(quincena_clave, dame_ultimo_dia=True)
    except ValueError:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, APOYOS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Consultar el concepto con clave PAZ, si no se encuentra, error
    concepto_paz = Concepto.query.filter_by(clave="PAZ").first()
    if concepto_paz is None:
        click.echo("ERROR: No existe el concepto con clave PAZ")
        sys.exit(1)

    # Consultar el concepto con clave DPL, si no se encuentra, error
    concepto_dpl = Concepto.query.filter_by(clave="DPL").first()
    if concepto_dpl is None:
        click.echo("ERROR: No existe el concepto con clave DPL")
        sys.exit(1)

    # Consultar el concepto con clave D62, si no se encuentra, error
    concepto_d62 = Concepto.query.filter_by(clave="D62").first()
    if concepto_d62 is None:
        click.echo("ERROR: No existe el concepto con clave D62")
        sys.exit(1)

    # Consultar el concepto con clave DPA, si no se encuentra, error
    concepto_dpa = Concepto.query.filter_by(clave="DPA").first()
    if concepto_dpa is None:
        click.echo("ERROR: No existe el concepto con clave DPA")
        sys.exit(1)

    # Consultar el concepto con clave DPJ, si no se encuentra, error
    concepto_dpj = Concepto.query.filter_by(clave="DPJ").first()
    if concepto_dpj is None:
        click.echo("ERROR: No existe el concepto con clave DPJ")
        sys.exit(1)

    # Consultar el concepto con clave DAZ, si no se encuentra, error
    concepto_daz = Concepto.query.filter_by(clave="DAZ").first()
    if concepto_daz is None:
        click.echo("ERROR: No existe el concepto con clave DAZ")
        sys.exit(1)

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    nominas_existentes = []
    personas_inexistentes = []
    plazas_inexistentes = []

    # Bucle por cada fila
    click.echo("Alimentando Nominas de Apoyos Anuales: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        rfc = hoja.cell_value(fila, 2).strip().upper()
        centro_trabajo_clave = hoja.cell_value(fila, 3).strip().upper()
        plaza_clave = hoja.cell_value(fila, 4).strip().upper()
        percepcion = float(hoja.cell_value(fila, 7))
        deduccion = float(hoja.cell_value(fila, 8))
        impte = float(hoja.cell_value(fila, 9))
        desde_s = str(int(hoja.cell_value(fila, 5)))
        hasta_s = str(int(hoja.cell_value(fila, 6)))
        fecha_pago_en_hoja = hoja.cell_value(fila, 10)  # Debe ser una fecha

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Si viene la fecha de pago en la hoja se toma, de lo contrario se usa el parámetro de la fecha de pago
        if isinstance(fecha_pago_en_hoja, datetime):
            fecha_pago = fecha_pago_en_hoja

        # Tomar el importe del concepto PAZ, si no esta presente sera cero
        try:
            impte_concepto_paz = float(hoja.cell_value(fila, 11))
        except ValueError:
            impte_concepto_paz = 0.0

        # Tomar el importe del concepto DPL, si no esta presente sera cero
        try:
            impte_concepto_dpl = float(hoja.cell_value(fila, 12))
        except ValueError:
            impte_concepto_dpl = 0.0

        # Tomar el importe del concepto D62, si no esta presente sera cero
        try:
            impte_concepto_d62 = float(hoja.cell_value(fila, 13))
        except ValueError:
            impte_concepto_d62 = 0.0

        # Tomar el importe del concepto DPA, si no esta presente sera cero
        try:
            impte_concepto_dpa = float(hoja.cell_value(fila, 14))
        except ValueError:
            impte_concepto_dpa = 0.0

        # Tomar el importe del concepto DPJ, si no esta presente sera cero
        try:
            impte_concepto_dpj = float(hoja.cell_value(fila, 15))
        except ValueError:
            impte_concepto_dpj = 0.0

        # Tomar el importe del concepto DAZ, si no esta presente sera cero
        try:
            impte_concepto_daz = float(hoja.cell_value(fila, 16))
        except ValueError:
            impte_concepto_daz = 0.0

        # Consultar la persona
        persona = Persona.query.filter_by(rfc=rfc).first()

        # Si NO existe, se agrega a la lista de personas_inexistentes y se salta
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()

        # Si NO existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Consultar la Plaza
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()

        # Si NO existe se agrega a la lista de plazas_inexistentes y se salta
        if plaza is None:
            plazas_inexistentes.append(plaza_clave)
            continue

        # Revisar que en nominas no exista una nomina con la misma persona, quincena y tipo APOYO ANUAL, si existe se omite
        nominas_posibles = (
            Nomina.query.filter_by(persona_id=persona.id)
            .filter_by(quincena_id=quincena.id)
            .filter_by(tipo="APOYO ANUAL")
            .filter_by(estatus="A")
            .all()
        )
        if len(nominas_posibles) > 0:
            nominas_existentes.append(rfc)
            continue

        # Constante tipo
        TIPO = "APOYO ANUAL"

        # Alimentar impte_concepto_paz, con concepto PAZ
        if impte_concepto_paz != 0.0:
            if probar is False:
                percepcion_deduccion_paz = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_paz,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_paz,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_paz)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_dpl, con concepto DPL
        if impte_concepto_dpl != 0.0:
            if probar is False:
                percepcion_deduccion_paz = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_dpl,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_dpl,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_paz)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar impte_concepto_d62, con concepto D62
        if impte_concepto_d62 != 0.0:
            if probar is False:
                percepcion_deduccion_d62 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d62,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d62,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d62)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar impte_concepto_dpa, con concepto DPA
        if impte_concepto_dpa != 0.0:
            if probar is False:
                percepcion_deduccion_dpa = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_dpa,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_dpa,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_dpa)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar impte_concepto_dpj, con concepto DPJ
        if impte_concepto_dpj != 0.0:
            if probar is False:
                percepcion_deduccion_dpj = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_dpj,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_dpj,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_dpj)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar impte_concepto_daz, con concepto DAZ
        if impte_concepto_daz != 0.0:
            if probar is False:
                percepcion_deduccion_daz = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_daz,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_daz,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_daz)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo=TIPO,
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style("a", fg="green"), nl=False)

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        # sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Actualizar la quincena para poner en verdadero el campo tiene_apoyos_anuales
    quincena.tiene_apoyos_anuales = True
    quincena.save()

    # Si hubo centros_trabajos_inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas_inexistentes, mostrarlos
    if len(plazas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(plazas_inexistentes)} Plazas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(plazas_inexistentes)}", fg="yellow"))

    # Si hubo personas_inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Si hubo nominas_existentes, mostrarlos
    if len(nominas_existentes) > 0:
        click.echo(click.style(f"  Hubo {len(nominas_existentes)} Apoyos Anuales que ya existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(nominas_existentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Alimentar Apoyos Anuales: {contador} insertadas en la quincena {quincena_clave}.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_aguinaldos_retroactivos(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar aguinaldos retroactivos"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Definir la fecha_final en base a la clave de la quincena
    try:
        fecha_final = quincena_to_fecha(quincena_clave, dame_ultimo_dia=True)
    except ValueError:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, RETROACTIVOS_AGUINALDOS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Consultar el concepto con clave PGA, si no se encuentra, error
    concepto_pga = Concepto.query.filter_by(clave="PGA").first()
    if concepto_pga is None:
        click.echo("ERROR: No existe el concepto con clave PGA")
        sys.exit(1)

    # Consultar el concepto con clave P22, si no se encuentra, error
    concepto_p22 = Concepto.query.filter_by(clave="P22").first()
    if concepto_p22 is None:
        click.echo("ERROR: No existe el concepto con clave P22")
        sys.exit(1)

    # Consultar el concepto con clave PGV, si no se encuentra, error
    concepto_pgv = Concepto.query.filter_by(clave="PGV").first()
    if concepto_pgv is None:
        click.echo("ERROR: No existe el concepto con clave PGV")
        sys.exit(1)

    # Consultar el concepto con clave P28, si no se encuentra, error
    concepto_p28 = Concepto.query.filter_by(clave="P28").first()
    if concepto_p28 is None:
        click.echo("ERROR: No existe el concepto con clave P28")
        sys.exit(1)

    # Consultar el concepto con clave D01, si no se encuentra, error
    concepto_d01 = Concepto.query.filter_by(clave="D01").first()
    if concepto_d01 is None:
        click.echo("ERROR: No existe el concepto con clave D01")
        sys.exit(1)

    # Consultar el concepto con clave D62, si no se encuentra, error
    concepto_d62 = Concepto.query.filter_by(clave="D62").first()
    if concepto_d62 is None:
        click.echo("ERROR: No existe el concepto con clave D62")
        sys.exit(1)

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    nominas_existentes = []
    personas_inexistentes = []
    plazas_inexistentes = []

    # Bucle por cada fila
    click.echo("Alimentando Nominas de Apoyos Anuales: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        # 0: QUINCENA
        # 1: TIPO
        rfc = hoja.cell_value(fila, 2).strip().upper()  # 2: RFC
        centro_trabajo_clave = hoja.cell_value(fila, 3).strip().upper()  # 3: CENTRO_TRABAJO
        plaza_clave = hoja.cell_value(fila, 4).strip().upper()  # 4: PLAZA
        desde_s = str(int(hoja.cell_value(fila, 5)))  # 5: DESDE
        hasta_s = str(int(hoja.cell_value(fila, 6)))  # 6: HASTA
        percepcion = int(hoja.cell_value(fila, 7))  # 7: PERCEPCION
        deduccion = int(hoja.cell_value(fila, 8))  # 8: DEDUCCION
        impte = int(hoja.cell_value(fila, 9))  # 9: IMPORTE
        fecha_pago_en_hoja = hoja.cell_value(fila, 10)  # Debe ser una fecha
        # 11: PGA
        # 12: P22
        # 13: PGV
        # 14: P28
        # 15: D01
        # 16: D62

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Si viene la fecha de pago en la hoja se toma, de lo contrario se usa el parámetro de la fecha de pago
        if isinstance(fecha_pago_en_hoja, datetime):
            fecha_pago = fecha_pago_en_hoja

        # Tomar el importe del concepto PGA, si no esta presente sera cero
        try:
            impte_concepto_pga = float(hoja.cell_value(fila, 11))
        except ValueError:
            impte_concepto_pga = 0.0

        # Tomar el importe del concepto P22, si no esta presente sera cero
        try:
            impte_concepto_p22 = float(hoja.cell_value(fila, 12))
        except ValueError:
            impte_concepto_p22 = 0.0

        # Tomar el importe del concepto PGV, si no esta presente sera cero
        try:
            impte_concepto_pgv = float(hoja.cell_value(fila, 13))
        except ValueError:
            impte_concepto_pgv = 0.0

        # Tomar el importe del concepto P28, si no esta presente sera cero
        try:
            impte_concepto_p28 = float(hoja.cell_value(fila, 14))
        except ValueError:
            impte_concepto_p28 = 0.0

        # Tomar el importe del concepto D01, si no esta presente sera cero
        try:
            impte_concepto_d01 = float(hoja.cell_value(fila, 15))
        except ValueError:
            impte_concepto_d01 = 0.0

        # Tomar el importe del concepto D62, si no esta presente sera cero
        try:
            impte_concepto_d62 = float(hoja.cell_value(fila, 16))
        except ValueError:
            impte_concepto_d62 = 0.0

        # Consultar la persona
        persona = Persona.query.filter_by(rfc=rfc).first()

        # Si NO existe, se agrega a la lista de personas_inexistentes y se salta
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()

        # Si NO existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Consultar la Plaza
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()

        # Si NO existe se agrega a la lista de plazas_inexistentes y se salta
        if plaza is None:
            plazas_inexistentes.append(plaza_clave)
            continue

        # Revisar que en nominas no exista una nomina con la misma persona, quincena y tipo RETROACTIVO AGUINALDO, si existe se omite
        nominas_posibles = (
            Nomina.query.filter_by(persona_id=persona.id)
            .filter_by(quincena_id=quincena.id)
            .filter_by(tipo="RETROACTIVO AGUINALDO")
            .filter_by(estatus="A")
            .all()
        )
        if len(nominas_posibles) > 0:
            nominas_existentes.append(rfc)
            continue

        # Constante tipo
        TIPO = "RETROACTIVO AGUINALDO"

        # Alimentar impte_concepto_pga, con concepto PGA
        if impte_concepto_pga != 0.0:
            if probar is False:
                percepcion_deduccion_pga = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pga,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_pga,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_pga)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_p22, con concepto P22
        if impte_concepto_p22 != 0.0:
            if probar is False:
                percepcion_deduccion_p22 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_p22,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_p22,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_p22)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_pgv, con concepto PGV
        if impte_concepto_pgv != 0.0:
            if probar is False:
                percepcion_deduccion_pgv = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pgv,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_pgv,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_pgv)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_p28, con concepto P28
        if impte_concepto_p28 != 0.0:
            if probar is False:
                percepcion_deduccion_p28 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_p28,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_p28,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_p28)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_d01, con concepto D01
        if impte_concepto_d01 != 0.0:
            if probar is False:
                percepcion_deduccion_d01 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d01,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d01,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d01)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar impte_concepto_d62, con concepto D62
        if impte_concepto_d62 != 0.0:
            if probar is False:
                percepcion_deduccion_d62 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d62,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d62,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d62)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo=TIPO,
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style("a", fg="green"), nl=False)

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Actualizar la quincena para poner en verdadero el campo tiene_apoyos_anuales
    quincena.tiene_apoyos_anuales = True
    quincena.save()

    # Si hubo centros_trabajos_inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas_inexistentes, mostrarlos
    if len(plazas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(plazas_inexistentes)} Plazas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(plazas_inexistentes)}", fg="yellow"))

    # Si hubo personas_inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Si hubo nominas_existentes, mostrarlos
    if len(nominas_existentes) > 0:
        click.echo(click.style(f"  Hubo {len(nominas_existentes)} Apoyos Anuales que ya existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(nominas_existentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Alimentar Apoyos Anuales: {contador} insertadas en la quincena {quincena_clave}.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_apoyos_madres(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar apoyos del dia de la madre"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Definir la fecha_final en base a la clave de la quincena
    try:
        fecha_final = quincena_to_fecha(quincena_clave, dame_ultimo_dia=True)
    except ValueError:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, APOYOS_DIA_DE_LA_MADRE_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Consultar el concepto con clave PA5, si no se encuentra, error
    concepto_pa5 = Concepto.query.filter_by(clave="PA5").first()
    if concepto_pa5 is None:
        click.echo("ERROR: No existe el concepto con clave PA5")
        sys.exit(1)

    # Consultar el concepto con clave DPL, si no se encuentra, error
    concepto_dpl = Concepto.query.filter_by(clave="DPL").first()
    if concepto_dpl is None:
        click.echo("ERROR: No existe el concepto con clave DPL")
        sys.exit(1)

    # Consultar el concepto con clave D62, si no se encuentra, error
    concepto_d62 = Concepto.query.filter_by(clave="D62").first()
    if concepto_d62 is None:
        click.echo("ERROR: No existe el concepto con clave D62")
        sys.exit(1)

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    nominas_existentes = []
    personas_inexistentes = []
    plazas_inexistentes = []

    # Bucle por cada fila
    click.echo("Alimentando Nominas de Apoyos Anuales: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        rfc = hoja.cell_value(fila, 2).strip().upper()
        centro_trabajo_clave = hoja.cell_value(fila, 3).strip().upper()
        plaza_clave = hoja.cell_value(fila, 4).strip().upper()
        percepcion = float(hoja.cell_value(fila, 7))
        deduccion = float(hoja.cell_value(fila, 8))
        impte = float(hoja.cell_value(fila, 9))
        desde_s = str(int(hoja.cell_value(fila, 5)))
        hasta_s = str(int(hoja.cell_value(fila, 6)))
        fecha_pago_en_hoja = hoja.cell_value(fila, 10)  # Debe ser una fecha

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Si viene la fecha de pago en la hoja se toma, de lo contrario se usa el parámetro de la fecha de pago
        if isinstance(fecha_pago_en_hoja, datetime):
            fecha_pago = fecha_pago_en_hoja

        # Tomar el importe del concepto PAZ, si no esta presente sera cero
        try:
            impte_concepto_pa5 = float(hoja.cell_value(fila, 11))
        except ValueError:
            impte_concepto_pa5 = 0.0

        # Tomar el importe del concepto DPL, si no esta presente sera cero
        try:
            impte_concepto_dpl = float(hoja.cell_value(fila, 12))
        except ValueError:
            impte_concepto_dpl = 0.0

        # Tomar el importe del concepto D62, si no esta presente sera cero
        try:
            impte_concepto_d62 = float(hoja.cell_value(fila, 13))
        except ValueError:
            impte_concepto_d62 = 0.0

        # Consultar la persona
        persona = Persona.query.filter_by(rfc=rfc).first()

        # Si NO existe, se agrega a la lista de personas_inexistentes y se salta
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()

        # Si NO existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Consultar la Plaza
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()

        # Si NO existe se agrega a la lista de plazas_inexistentes y se salta
        if plaza is None:
            plazas_inexistentes.append(plaza_clave)
            continue

        # Revisar que en nominas no exista una nomina con la misma persona, quincena y tipo APOYO DIA DE LA MADRE, si existe se omite
        nominas_posibles = (
            Nomina.query.filter_by(persona_id=persona.id)
            .filter_by(quincena_id=quincena.id)
            .filter_by(tipo="APOYO DIA DE LA MADRE")
            .filter_by(estatus="A")
            .all()
        )
        if len(nominas_posibles) > 0:
            nominas_existentes.append(rfc)
            continue

        # Constante tipo
        TIPO = "APOYO DIA DE LA MADRE"

        # Alimentar impte_concepto_paz, con concepto PA5
        if impte_concepto_pa5 != 0.0:
            if probar is False:
                percepcion_deduccion_paz = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pa5,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_pa5,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_paz)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_dpl, con concepto DPL
        if impte_concepto_dpl != 0.0:
            if probar is False:
                percepcion_deduccion_paz = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_dpl,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_dpl,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_paz)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar impte_concepto_d62, con concepto D62
        if impte_concepto_d62 != 0.0:
            if probar is False:
                percepcion_deduccion_d62 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d62,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d62,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d62)
            click.echo(click.style("d", fg="blue"), nl=False)

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo=TIPO,
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style("a", fg="green"), nl=False)

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Actualizar la quincena para poner en verdadero el campo tiene_apoyos_anuales
    quincena.tiene_apoyos_anuales = True
    quincena.save()

    # Si hubo centros_trabajos_inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas_inexistentes, mostrarlos
    if len(plazas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(plazas_inexistentes)} Plazas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(plazas_inexistentes)}", fg="yellow"))

    # Si hubo personas_inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Si hubo nominas_existentes, mostrarlos
    if len(nominas_existentes) > 0:
        click.echo(click.style(f"  Hubo {len(nominas_existentes)} Apoyos Anuales que ya existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(nominas_existentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Alimentar Apoyos Anuales: {contador} insertadas en la quincena {quincena_clave}.", fg="green"))


@click.command()
@click.argument("archivo_xlsx", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_extraordinarios(archivo_xlsx: str, probar: bool = False):
    """Alimentar extraordinarios"""

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos
    if EXTRAORDINARIOS_BASE_DIR is None:
        click.echo("ERROR: Variable de entorno EXTRAORDINARIOS_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXTRAORDINARIOS_BASE_DIR, archivo_xlsx)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # El archivo debe tener los siguientes encabezados
    # 01: RFC
    # 02: QUINCENA
    # 03: CLAVE CENTRO TRABAJO
    # 04: PLAZA	DESDE
    # 05: HASTA
    # 06: TIPO NOMINA
    # 07: PERCEPCION
    # 08: DEDUCCION
    # 09: IMPORTE
    # 10: NUM CHEQUE
    # 11: FECHA DE PAGO
    # 12: TIPO DE EXTRAORDINARIA
    # 13: P30
    # 14: PGN
    # 15: PGA
    # 16: P22
    # 17: PVD
    # 18: PGP
    # 19: P20
    # 20: PAM
    # 21: PS3
    # 22: D01
    # 23: D1A
    # 24: P07
    # 25: P7G
    # 26: PHR
    # 27: PFB

    # Definir el diccionario con los numeros de columnas y las claves de los conceptos
    numeros_columnas_conceptos_claves = {
        13: "P30",
        14: "PGN",
        15: "PGA",
        16: "P22",
        17: "PVD",
        18: "PGP",
        19: "P20",
        20: "PAM",
        21: "PS3",
        22: "D01",
        23: "D1A",
        24: "P07",
        25: "P7G",
        26: "PHR",
        27: "PFB",
    }

    # Consultar el diccionario con los numeros de columnas y las consultas a los conceptos
    click.echo("Consultando conceptos...")
    numeros_columnas_conceptos = {}
    for num_col, concepto_clave in numeros_columnas_conceptos_claves.items():
        concepto = Concepto.query.filter_by(clave=concepto_clave).first()
        if concepto is None:
            click.echo(click.style(f"  ADVERTENCIA: No existe el concepto con clave {concepto_clave}", fg="yellow"))
        numeros_columnas_conceptos[num_col] = concepto

    # Leer archivo_xlsx con openpyxl
    workbook = load_workbook(filename=ruta, read_only=True, data_only=True)

    # Inicializar listado de personas NO encontradas
    personas_no_encontradas = []

    # Inicializar listado de quincenas NO encontradas
    quincenas_no_encontradas = []

    # Inicializar desde no validos
    desde_no_validos = []

    # Inicializar hasta no validos
    hasta_no_validos = []

    # Bucle por los renglones de la hoja
    contador = 0
    click.echo("Alimentando Extraordinarios:")
    for row in workbook.active.iter_rows(min_row=2, max_col=28, max_row=3000):
        # Juntar todas las celdas de la fila en una lista
        fila = [celda.value for celda in row]

        # Si el primer elemento de la fila es vacio, se rompe el ciclo
        if str(fila[0]).strip() == "" or fila[0] is None or fila[0] == "None":
            break

        # Tomar los valores de la fila
        rfc = safe_rfc(fila[0])
        quincena_clave = safe_quincena(str(fila[1]))
        centro_trabajo_clave = safe_clave(fila[2], max_len=10)
        plaza_clave = safe_clave(fila[3], max_len=24)
        desde_dt = fila[4]
        hasta_dt = fila[5]
        # tipo_nomina = fila[6]
        percepcion = fila[7]
        deduccion = fila[8]
        importe = fila[9]
        # num_cheque = fila[10]
        fecha_pago = fila[11]
        # tipo_extraordinaria = fila[12]

        # Consultar el centro de trabajo a partir de la clave, si no se encuentra, se usa el ND
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()
        if centro_trabajo is None:
            centro_trabajo = CentroTrabajo.query.filter_by(clave="ND").first()

        # Consultar la persona a partir del rfc, si no se encuentra, se omite
        persona = Persona.query.filter_by(rfc=rfc).first()
        if persona is None:
            personas_no_encontradas.append(rfc)
            click.echo(click.style(f"[{rfc}]", fg="yellow"))
            continue

        # Consultar la plaza a partir de la clave, si no se encuentra, se usa el ND
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()
        if plaza is None:
            plaza = Plaza.query.filter_by(clave="ND").first()

        # Validar quincena
        if re.match(QUINCENA_REGEXP, quincena_clave) is None:
            quincenas_no_encontradas.append(quincena_clave)
            click.echo(click.style("f[{quincena_clave}]", fg="yellow"))
            continue

        # Consultar la quincena a partir de la clave, si no se encuentra, se omite
        quincena = Quincena.query.filter_by(clave=quincena_clave).first()
        if quincena is None:
            quincenas_no_encontradas.append(quincena_clave)
            click.echo(click.style(f"[{quincena_clave}]", fg="yellow"))
            continue

        # Validar desde
        try:
            desde_clave = crear_clave_quincena(desde_dt.date())
            desde_dt = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
        except ValueError:
            desde_no_validos.append(desde_dt)
            click.echo(click.style("[desde]", fg="yellow"))
            continue

        # Validar hasta
        try:
            hasta_clave = crear_clave_quincena(hasta_dt.date())
            hasta_dt = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            hasta_no_validos.append(hasta_dt)
            click.echo(click.style("[hasta]", fg="yellow"))
            continue

        # Si probar es falso
        if probar is False:
            # Alimentar nomina
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde_dt,
                desde_clave=desde_clave,
                hasta=hasta_dt,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=importe,
                tipo="EXTRAORDINARIO",
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style(f"  {rfc} {quincena_clave}: ", fg="cyan"), nl=False)

        # Tomar los valores de la hoja de calculo y alimentar las percepciones y deducciones
        for num_col, concepto in numeros_columnas_conceptos.items():
            valor = fila[num_col]
            if concepto is None:
                click.echo(click.style("X", fg="yellow"), nl=False)
                continue
            if valor is None:
                click.echo(click.style("_", fg="yellow"), nl=False)
                continue
            if float(valor) == 0:
                click.echo(click.style("0", fg="yellow"), nl=False)
                continue
            if probar is False:
                percepcion_deduccion = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=valor,
                    tipo="EXTRAORDINARIO",
                )
                sesion.add(percepcion_deduccion)
            click.echo(click.style("+", fg="green"), nl=False)

        # Incrementar contador
        contador += 1
        click.echo("")

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Si hubo personas_no_encontradas, mostrarlos
    if len(personas_no_encontradas) > 0:
        click.echo(click.style(f"  Hubo {len(personas_no_encontradas)} Personas no encontradas.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_no_encontradas)}", fg="yellow"))

    # Si hubo quincenas_no_encontradas, mostrarlos
    if len(quincenas_no_encontradas) > 0:
        click.echo(click.style(f"  Hubo {len(quincenas_no_encontradas)} Quincenas no encontradas.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(quincenas_no_encontradas)}", fg="yellow"))

    # Si hubo desde_no_validos, mostrarlos
    if len(desde_no_validos) > 0:
        click.echo(click.style(f"  Hubo {len(desde_no_validos)} Desde no validos.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(desde_no_validos)}", fg="yellow"))

    # Si hubo hasta_no_validos, mostrarlos
    if len(hasta_no_validos) > 0:
        click.echo(click.style(f"  Hubo {len(hasta_no_validos)} Hasta no validos.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(hasta_no_validos)}", fg="yellow"))

    # Mensaje termino
    if probar:
        click.echo(click.style(f"Alimentar Extraordinarios: modo PROBAR {contador} pueden insertarse.", fg="green"))
    else:
        click.echo(click.style(f"Alimentar Extraordinarios: {contador} insertados.", fg="green"))


@click.command()
@click.argument("archivo_xlsx", type=str)
@click.argument("quincena_clave", type=str)
@click.option("--tipo", type=click.Choice(["AGUINALDO", "APOYO ANUAL", "APOYO DIA DE LA MADRE", "SALARIO"]), default="SALARIO")
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_pensiones_alimenticias(
    archivo_xlsx: str,
    quincena_clave: str,
    tipo: str,
    probar: bool = True,
):
    """Alimentar pensiones alimenticias"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar el directorio donde espera encontrar los archivos
    if PENSIONES_ALIMENTICIAS_BASE_DIR is None:
        click.echo("ERROR: Variable de entorno PENSIONES_ALIMENTICIAS_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(PENSIONES_ALIMENTICIAS_BASE_DIR, archivo_xlsx)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar la quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()
    if quincena is None:
        click.echo("ERROR: Quincena no encontrada.")
        sys.exit(1)

    # Consultar el concepto con clave P62, si no se encuentra, error
    concepto_p62 = Concepto.query.filter_by(clave="P62").first()
    if concepto_p62 is None:
        click.echo("ERROR: No existe el concepto con clave P62")
        sys.exit(1)

    # Consultar el concepto con clave P62A, si no se encuentra, error
    concepto_p62a = Concepto.query.filter_by(clave="P62A").first()
    if concepto_p62a is None:
        click.echo("ERROR: No existe el concepto con clave P62A")
        sys.exit(1)

    # Consultar el concepto con clave P62B, si no se encuentra, error
    concepto_p62b = Concepto.query.filter_by(clave="P62B").first()
    if concepto_p62b is None:
        click.echo("ERROR: No existe el concepto con clave P62B")
        sys.exit(1)

    # Consultar el concepto con clave PA5, si no se encuentra, error
    concepto_pa5 = Concepto.query.filter_by(clave="PA5").first()
    if concepto_pa5 is None:
        click.echo("ERROR: No existe el concepto con clave PA5")
        sys.exit(1)

    # De acuerdo al TIPO es el concepto que se usara
    conceptos = {
        "SALARIO": concepto_p62,
        "AGUINALDO": concepto_p62a,
        "APOYO ANUAL": concepto_p62b,
        "APOYO DIA DE LA MADRE": concepto_pa5,
    }

    # El archivo debe tener los siguientes encabezados
    # 00: RFC
    # 01: QUINCENA
    # 02: CENTRO DE TRABAJO
    # 03: PLAZA
    # 04: DESDE
    # 05: HASTA
    # 06: TIPO
    # 07: PERCEPCION
    # 08: DEDUCCION
    # 09: IMPORTE
    # 10: NO. CHEQUE
    # 11: FECHA DE PAGO
    # 12: CONCEPTO (P62 si TIPO es SALARIO, P62A si es AGUINALDO, P62B si es APOYO ANUAL)

    # Leer archivo_xlsx con openpyxl
    workbook = load_workbook(filename=ruta, read_only=True, data_only=True)

    # Inicializar listado de personas NO encontradas
    personas_no_encontradas = []

    # Inicializar desde no validos
    desde_no_validos = []

    # Inicializar hasta no validos
    hasta_no_validos = []

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Bucle por los renglones de la hoja
    contador = 0
    click.echo("Alimentando Pensiones Alimenticias:")
    for row in workbook.active.iter_rows(min_row=2, max_col=16, max_row=3000):
        # Juntar todas las celdas de la fila en una lista
        fila = [celda.value for celda in row]

        # Si el primer elemento de la fila es vacio, se rompe el ciclo
        if str(fila[0]).strip() == "" or fila[0] is None or fila[0] == "None":
            break

        # Tomar los valores de la fila
        rfc = safe_rfc(fila[0])
        centro_trabajo_clave = safe_clave(fila[2], max_len=10)
        plaza_clave = safe_clave(fila[3], max_len=24)
        desde_clave = str(fila[4])
        hasta_clave = str(fila[5])
        percepcion = fila[7]
        deduccion = fila[8]
        importe = fila[9]
        num_cheque = fila[10]
        fecha_pago = fila[11]
        importe_concepto = fila[12]

        # Consultar la persona a partir del rfc, si no se encuentra, se omite
        persona = Persona.query.filter_by(rfc=rfc).first()
        if persona is None:
            personas_no_encontradas.append(rfc)
            continue

        # Consultar el centro de trabajo a partir de la clave, si no se encuentra, se usa el ND
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()
        if centro_trabajo is None:
            centro_trabajo = CentroTrabajo.query.filter_by(clave="ND").first()

        # Consultar la plaza a partir de la clave, si no se encuentra, se usa el ND
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()
        if plaza is None:
            plaza = Plaza.query.filter_by(clave="ND").first()

        # Validar desde
        try:
            desde_dt = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
        except ValueError:
            desde_no_validos.append(desde_clave)
            continue

        # Validar hasta
        try:
            hasta_dt = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            hasta_no_validos.append(hasta_clave)
            continue

        # Si tiene importe_concepto, alimentar registro en PercepcionDeduccion
        if importe_concepto > 0:
            concepto = conceptos[tipo]
            if probar is False:
                percepcion_deduccion = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=importe_concepto,
                    tipo=tipo,
                )
                sesion.add(percepcion_deduccion)
                click.echo(click.style("p", fg="green"), nl=False)
            else:
                click.echo(click.style(f"{concepto.clave}: {importe_concepto}, ", fg="green"), nl=False)

        # Si probar es falso
        if probar is False:
            # Alimentar nomina
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde_dt,
                desde_clave=desde_clave,
                hasta=hasta_dt,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=importe,
                tipo="PENSION ALIMENTICIA",
                fecha_pago=fecha_pago,
                num_cheque=num_cheque,
            )
            sesion.add(nomina)
            click.echo(click.style("n", fg="cyan"), nl=False)
        else:
            click.echo(click.style(f"{rfc}: {importe}, ", fg="cyan"), nl=False)

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Si hubo personas_no_encontradas, mostrarlos
    if len(personas_no_encontradas) > 0:
        click.echo(click.style(f"  Hubo {len(personas_no_encontradas)} Personas no encontradas.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_no_encontradas)}", fg="yellow"))

    # Si hubo desde_no_validos, mostrarlos
    if len(desde_no_validos) > 0:
        click.echo(click.style(f"  Hubo {len(desde_no_validos)} Desde no validos.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(desde_no_validos)}", fg="yellow"))

    # Si hubo hasta_no_validos, mostrarlos
    if len(hasta_no_validos) > 0:
        click.echo(click.style(f"  Hubo {len(hasta_no_validos)} Hasta no validos.", fg="yellow"))
        click.echo(click.style(f"  {', '.join(hasta_no_validos)}", fg="yellow"))

    # Mensaje termino
    if probar:
        click.echo(click.style(f"Alimentar Pensiones Alimenticias: modo PROBAR {contador} pueden insertarse.", fg="green"))
    else:
        click.echo(click.style(f"Alimentar Pensiones Alimenticias: {contador} insertados.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_primas_vacacionales(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar primas vacacionales"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Definir la fecha_final con base en la clave de la quincena
    try:
        fecha_final = quincena_to_fecha(quincena_clave, dame_ultimo_dia=True)
    except ValueError:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, PRIMAS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = sesion.query(Quincena).filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Consultar el concepto con clave P20, si no se encuentra, error
    concepto_p20 = sesion.query(Concepto).filter_by(clave="P20").first()
    if concepto_p20 is None:
        click.echo("ERROR: No existe el concepto con clave P20")
        sys.exit(1)

    # Consultar el concepto con clave PGP, si no se encuentra, error
    concepto_pgp = sesion.query(Concepto).filter_by(clave="PGP").first()
    if concepto_pgp is None:
        click.echo("ERROR: No existe el concepto con clave PGP")
        sys.exit(1)

    # Consultar el concepto con clave PGV, si no se encuentra, error
    concepto_pgv = sesion.query(Concepto).filter_by(clave="PGV").first()
    if concepto_pgv is None:
        click.echo("ERROR: No existe el concepto con clave PGV")
        sys.exit(1)

    # Consultar el concepto con clave D1R, si no se encuentra, error
    concepto_d1r = sesion.query(Concepto).filter_by(clave="D1R").first()
    if concepto_d1r is None:
        click.echo("ERROR: No existe el concepto con clave D1R")
        sys.exit(1)

    # Consultar el concepto con clave D62, si no se encuentra, error
    concepto_d62 = sesion.query(Concepto).filter_by(clave="D62").first()
    if concepto_d62 is None:
        click.echo("ERROR: No existe el concepto con clave D62")
        sys.exit(1)

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    nominas_existentes = []
    personas_inexistentes = []
    plazas_inexistentes = []

    # Bucle por cada fila
    click.echo("Alimentando Nominas de Primas: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        # Columna 00 UNIDAD_DIS
        centro_trabajo_clave = hoja.cell_value(fila, 1).strip().upper()  # Columna 01 CLAVE_CT
        rfc = hoja.cell_value(fila, 2).strip().upper()  # Columna 02 RFC
        # Columna 03 NOMBRE
        # Columna 04 MUNICIPIO
        # Columna 05 ZONA_ESCO
        # Columna 06 LOCALIDAD
        # Columna 07 SOB_SUELDO
        plaza_clave = hoja.cell_value(fila, 8).strip().upper()  # Columna 08 PLAZA
        # Columna 09 NIVP
        # Columna 10 PROG_SUBP
        # Columna 11 URESP_PLA
        percepcion = int(hoja.cell_value(fila, 12)) / 100.0  # Columna 12 PERCEPCION
        deduccion = int(hoja.cell_value(fila, 13)) / 100.0  # Columna 13 DEDUCCION
        impte = int(hoja.cell_value(fila, 14)) / 100.0  # Columna 14 IMPTE
        # Columna 15 NO_CHEQUE
        desde_s = str(int(hoja.cell_value(fila, 16)))  # Columna 16 DESDE_S
        hasta_s = str(int(hoja.cell_value(fila, 17)))  # Columna 17 HASTA_S

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Consultar la persona
        persona = sesion.query(Persona).filter_by(rfc=rfc).first()

        # Si NO existe, se agrega a la lista de personas_inexistentes y se salta
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo
        centro_trabajo = sesion.query(CentroTrabajo).filter_by(clave=centro_trabajo_clave).first()

        # Si NO existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Consultar la Plaza
        plaza = sesion.query(Plaza).filter_by(clave=plaza_clave).first()

        # Si NO existe se agrega a la lista de plazas_inexistentes y se salta
        if plaza is None:
            plazas_inexistentes.append(plaza_clave)
            continue

        # Revisar que en nominas no exista una nomina con la misma persona, quincena y tipo APOYO ANUAL, si existe se omite
        # nominas_posibles = (
        #     sesion.query(Nomina).filter_by(persona_id=persona.id)
        #     .filter_by(quincena_id=quincena.id)
        #     .filter_by(tipo="PRIMA VACACIONAL")
        #     .filter_by(estatus="A")
        #     .all()
        # )
        # if len(nominas_posibles) > 0:
        #     nominas_existentes.append(rfc)
        #     continue

        # Inicializar los importes de las percepciones y deducciones
        impt_concepto_p20 = 0.0
        impt_concepto_pgp = 0.0
        impt_concepto_pgv = 0.0
        impt_concepto_d1r = 0.0
        impt_concepto_d62 = 0.0

        # Buscar percepciones y deducciones
        col_num = 26
        while True:
            # Tomar 'P' o 'D', primero
            p_o_d = safe_string(hoja.cell_value(fila, col_num))

            # Si 'P' o 'D' es un texto vacio, se rompe el ciclo
            if p_o_d == "":
                break

            # Tomar los dos caracteres adicionales del concepto
            conc = safe_string(hoja.cell_value(fila, col_num + 1))

            # Definir la clave del concepto
            concepto_clave = f"{p_o_d}{conc}"

            # Si la clave del concepto NO es P20, PGP, PGV, D1R o D62, se omite
            if concepto_clave not in ["P20", "PGP", "PGV", "D1R", "D62"]:
                click.echo(click.style("X", fg="yellow"), nl=False)
                col_num += 6
                continue

            # Tomar el importe
            try:
                impt = int(hoja.cell_value(fila, col_num + 3)) / 100.0
            except ValueError:
                impt = 0.0

            # Asignar el importe al concepto correspondiente
            if concepto_clave == "P20":
                impt_concepto_p20 = impt
            elif concepto_clave == "PGP":
                impt_concepto_pgp = impt
            elif concepto_clave == "PGV":
                impt_concepto_pgv = impt
            elif concepto_clave == "D1R":
                impt_concepto_d1r = impt
            elif concepto_clave == "D62":
                impt_concepto_d62 = impt

            # Incrementar col_num en SEIS
            col_num += 6

            # Romper el ciclo cuando se acaban las columnas a leer
            if col_num > 236:
                break

        # Alimentar percepcion en PercepcionDeduccion, con concepto P20
        if impt_concepto_p20 > 0:
            if probar is False:
                percepcion_deduccion_p20 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_p20,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impt_concepto_p20,
                    tipo="PRIMA VACACIONAL",
                )
                sesion.add(percepcion_deduccion_p20)
            click.echo(click.style("[P20]", fg="blue"), nl=False)

        # Alimentar percepcion en PercepcionDeduccion, con concepto PGP
        if impt_concepto_pgp > 0:
            if probar is False:
                percepcion_deduccion_pgp = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pgp,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impt_concepto_pgp,
                    tipo="PRIMA VACACIONAL",
                )
                sesion.add(percepcion_deduccion_pgp)
            click.echo(click.style("[PGP]", fg="blue"), nl=False)

        # Alimentar percepcion en PercepcionDeduccion, con concepto PGV
        if impt_concepto_pgv > 0:
            if probar is False:
                percepcion_deduccion_pgv = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pgv,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impt_concepto_pgv,
                    tipo="PRIMA VACACIONAL",
                )
                sesion.add(percepcion_deduccion_pgv)
            click.echo(click.style("[PGV]", fg="blue"), nl=False)

        # Alimentar percepcion en PercepcionDeduccion, con concepto D1R
        if impt_concepto_d1r > 0:
            if probar is False:
                percepcion_deduccion_d1r = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d1r,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impt_concepto_d1r,
                    tipo="PRIMA VACACIONAL",
                )
                sesion.add(percepcion_deduccion_d1r)
            click.echo(click.style("[D1R]", fg="blue"), nl=False)

        # Alimentar percepcion en PercepcionDeduccion, con concepto D62
        if impt_concepto_d62 > 0:
            if probar is False:
                percepcion_deduccion_d62 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d62,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impt_concepto_d62,
                    tipo="PRIMA VACACIONAL",
                )
                sesion.add(percepcion_deduccion_d62)
            click.echo(click.style("[D62]", fg="blue"), nl=False)

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo="PRIMA VACACIONAL",
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style("+", fg="green"), nl=False)

        # Cerrar la sesion para que se guarden todos los datos en la base de datos
        sesion.commit()
        sesion.close()

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Actualizar la quincena para poner en verdadero el campo tiene_apoyos_anuales
    quincena.tiene_primas = True
    quincena.save()

    # Si hubo centros_trabajos_inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas_inexistentes, mostrarlos
    if len(plazas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(plazas_inexistentes)} Plazas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(plazas_inexistentes)}", fg="yellow"))

    # Si hubo personas_inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Si hubo nominas_existentes, mostrarlos
    if len(nominas_existentes) > 0:
        click.echo(click.style(f"  Hubo {len(nominas_existentes)} Primas que ya existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(nominas_existentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(
        click.style(f"  Alimentar Primas Vacacionales: {contador} insertadas en la quincena {quincena_clave}.", fg="green")
    )


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_salarios_rl(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar salarios RL"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, SALARIOS_RL_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Consultar el concepto con clave P01, si no se encuentra, error
    concepto_p01 = Concepto.query.filter_by(clave="P01").first()
    if concepto_p01 is None:
        click.echo("ERROR: No existe el concepto con clave P01")
        sys.exit(1)

    # Consultar el concepto con clave P03, si no se encuentra, error
    concepto_p03 = Concepto.query.filter_by(clave="P03").first()
    if concepto_p03 is None:
        click.echo("ERROR: No existe el concepto con clave P03")
        sys.exit(1)

    # Consultar el concepto con clave D01, si no se encuentra, error
    concepto_d01 = Concepto.query.filter_by(clave="D01").first()
    if concepto_d01 is None:
        click.echo("ERROR: No existe el concepto con clave D01")
        sys.exit(1)

    # Consultar el concepto con clave D02, si no se encuentra, error
    concepto_d02 = Concepto.query.filter_by(clave="D02").first()
    if concepto_d02 is None:
        click.echo("ERROR: No existe el concepto con clave D02")
        sys.exit(1)

    # Consultar el concepto con clave D03, si no se encuentra, error
    concepto_d03 = Concepto.query.filter_by(clave="D03").first()
    if concepto_d03 is None:
        click.echo("ERROR: No existe el concepto con clave D03")
        sys.exit(1)

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    nominas_existentes = []
    personas_inexistentes = []
    plazas_inexistentes = []

    # Bucle por cada fila
    click.echo("Alimentando Nominas de Salarios RL: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        rfc = hoja.cell_value(fila, 2).strip().upper()
        centro_trabajo_clave = hoja.cell_value(fila, 3).strip().upper()
        plaza_clave = hoja.cell_value(fila, 4).strip().upper()
        percepcion = float(hoja.cell_value(fila, 7))
        deduccion = float(hoja.cell_value(fila, 8))
        impte = float(hoja.cell_value(fila, 9))
        desde_s = str(int(hoja.cell_value(fila, 5)))
        hasta_s = str(int(hoja.cell_value(fila, 6)))
        fecha_pago_en_hoja = hoja.cell_value(fila, 10)  # Debe ser una fecha

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Si viene la fecha de pago en la hoja se toma, de lo contrario se usa el parámetro de la fecha de pago
        if isinstance(fecha_pago_en_hoja, datetime):
            fecha_pago = fecha_pago_en_hoja

        # Tomar el importe del concepto P01, si no esta presente sera cero
        try:
            impte_concepto_p01 = float(hoja.cell_value(fila, 11))
        except ValueError:
            impte_concepto_p01 = 0.0

        # Tomar el importe del concepto P03, si no esta presente sera cero
        try:
            impte_concepto_p03 = float(hoja.cell_value(fila, 12))
        except ValueError:
            impte_concepto_p03 = 0.0

        # Tomar el importe del concepto D01, si no esta presente sera cero
        try:
            impte_concepto_d01 = float(hoja.cell_value(fila, 13))
        except ValueError:
            impte_concepto_d01 = 0.0

        # Tomar el importe del concepto D02, si no esta presente sera cero
        try:
            impte_concepto_d02 = float(hoja.cell_value(fila, 14))
        except ValueError:
            impte_concepto_d02 = 0.0

        # Tomar el importe del concepto D03, si no esta presente sera cero
        try:
            impte_concepto_d03 = float(hoja.cell_value(fila, 15))
        except ValueError:
            impte_concepto_d03 = 0.0

        # Consultar la persona
        persona = Persona.query.filter_by(rfc=rfc).first()

        # Si NO existe, se agrega a la lista de personas_inexistentes y se salta
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()

        # Si NO existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Consultar la Plaza
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()

        # Si NO existe se agrega a la lista de plazas_inexistentes y se salta
        if plaza is None:
            plazas_inexistentes.append(plaza_clave)
            continue

        # Revisar que en nominas no exista una nomina con la misma persona, quincena y tipo SALARIO, si existe se omite
        nominas_posibles = (
            Nomina.query.filter_by(persona_id=persona.id)
            .filter_by(quincena_id=quincena.id)
            .filter_by(tipo="SALARIO")
            .filter_by(estatus="A")
            .all()
        )
        if len(nominas_posibles) > 0:
            nominas_existentes.append(rfc)
            continue

        # Constante tipo
        TIPO = "SALARIO"

        # Alimentar impte_concepto_p01, con concepto P01
        if impte_concepto_p01 != 0.0:
            if probar is False:
                percepcion_deduccion_p01 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_p01,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_p01,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_p01)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_p03, con concepto P03
        if impte_concepto_p03 != 0.0:
            if probar is False:
                percepcion_deduccion_p03 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_p03,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_p03,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_p03)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_d01, con concepto D01
        if impte_concepto_d01 != 0.0:
            if probar is False:
                percepcion_deduccion_d01 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d01,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d01,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d01)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_d02, con concepto D02
        if impte_concepto_d02 != 0.0:
            if probar is False:
                percepcion_deduccion_d02 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d02,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d02,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d02)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_d03, con concepto D03
        if impte_concepto_d03 != 0.0:
            if probar is False:
                percepcion_deduccion_d03 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d03,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d03,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d03)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo=TIPO,
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style("a", fg="green"), nl=False)

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Actualizar la quincena para poner en verdadero el campo tiene_primas_vacacionales
    quincena.tiene_primas_vacacionales = True
    quincena.save()

    # Si hubo centros_trabajos_inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas_inexistentes, mostrarlos
    if len(plazas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(plazas_inexistentes)} Plazas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(plazas_inexistentes)}", fg="yellow"))

    # Si hubo personas_inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Si hubo nominas_existentes, mostrarlos
    if len(nominas_existentes) > 0:
        click.echo(click.style(f"  Hubo {len(nominas_existentes)} Salarios RL que ya existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(nominas_existentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(click.style(f"  Alimentar Salarios RL: {contador} insertadas en la quincena {quincena_clave}.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("fecha_pago_str", type=str)
@click.option("--probar", is_flag=True, help="Solo probar la lectura del archivo.")
def alimentar_primas_vacacionales_alterno(quincena_clave: str, fecha_pago_str: str, probar: bool = False):
    """Alimentar primas vacacionales alterno"""

    # Validar quincena
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo("ERROR: Quincena inválida.")
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d")
    except ValueError:
        click.echo("ERROR: Fecha de pago inválida")
        sys.exit(1)

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Validar el directorio donde espera encontrar los archivos de explotacion
    if EXPLOTACION_BASE_DIR == "":
        click.echo("ERROR: Variable de entorno EXPLOTACION_BASE_DIR no definida.")
        sys.exit(1)

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, PRIMAS_FILENAME_XLS)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si existe la quincena, pero no esta ABIERTA, entonces se termina
    if quincena and quincena.estado != "ABIERTA":
        click.echo(f"ERROR: Quincena {quincena_clave} no esta ABIERTA.")
        sys.exit(1)

    # Si existe la quincena, pero ha sido eliminada, entonces se termina
    if quincena and quincena.estatus != "A":
        click.echo(f"ERROR: Quincena {quincena_clave} esta sido eliminada.")
        sys.exit(1)

    # Si no existe la quincena, se agrega
    if quincena is None:
        quincena = Quincena(clave=quincena_clave, estado="ABIERTA")
        sesion.add(quincena)
        sesion.commit()

    # Consultar el concepto con clave P20, si no se encuentra, error
    concepto_p20 = Concepto.query.filter_by(clave="P20").first()
    if concepto_p20 is None:
        click.echo("ERROR: No existe el concepto con clave P20")
        sys.exit(1)

    # Consultar el concepto con clave PGP, si no se encuentra, error
    concepto_pgp = Concepto.query.filter_by(clave="PGP").first()
    if concepto_pgp is None:
        click.echo("ERROR: No existe el concepto con clave PGP")
        sys.exit(1)

    # Consultar el concepto con clave PGV, si no se encuentra, error
    concepto_pgv = Concepto.query.filter_by(clave="PGV").first()
    if concepto_pgv is None:
        click.echo("ERROR: No existe el concepto con clave PGV")
        sys.exit(1)

    # Consultar el concepto con clave D1R, si no se encuentra, error
    concepto_d1r = Concepto.query.filter_by(clave="D1R").first()
    if concepto_d1r is None:
        click.echo("ERROR: No existe el concepto con clave D1R")
        sys.exit(1)

    # Consultar el concepto con clave D62, si no se encuentra, error
    concepto_d62 = Concepto.query.filter_by(clave="D62").first()
    if concepto_d62 is None:
        click.echo("ERROR: No existe el concepto con clave D62")
        sys.exit(1)

    # Abrir el archivo XLS con xlrd
    libro = xlrd.open_workbook(str(ruta))

    # Obtener la primera hoja
    hoja = libro.sheet_by_index(0)

    # Iniciar contadores
    contador = 0
    centros_trabajos_inexistentes = []
    nominas_existentes = []
    personas_inexistentes = []
    plazas_inexistentes = []

    # Bucle por cada fila
    click.echo("Alimentando Nominas de Prima Vacacional: ", nl=False)
    for fila in range(1, hoja.nrows):
        # Tomar las columnas
        rfc = hoja.cell_value(fila, 2).strip().upper()
        centro_trabajo_clave = hoja.cell_value(fila, 3).strip().upper()
        plaza_clave = hoja.cell_value(fila, 4).strip().upper()
        percepcion = float(hoja.cell_value(fila, 7))
        deduccion = float(hoja.cell_value(fila, 8))
        impte = float(hoja.cell_value(fila, 9))
        desde_s = str(int(hoja.cell_value(fila, 5)))
        hasta_s = str(int(hoja.cell_value(fila, 6)))
        fecha_pago_en_hoja = hoja.cell_value(fila, 10)  # Debe ser una fecha

        # Validar desde y hasta
        try:
            desde_clave = safe_quincena(desde_s)
            desde = quincena_to_fecha(desde_clave, dame_ultimo_dia=False)
            hasta_clave = safe_quincena(hasta_s)
            hasta = quincena_to_fecha(hasta_clave, dame_ultimo_dia=True)
        except ValueError:
            click.echo(click.style(f"ERROR: Quincena inválida en '{desde_s}' o '{hasta_s}'", fg="red"))
            sys.exit(1)

        # Si viene la fecha de pago en la hoja se toma, de lo contrario se usa el parámetro de la fecha de pago
        if isinstance(fecha_pago_en_hoja, datetime):
            fecha_pago = fecha_pago_en_hoja

        # Tomar el importe del concepto P20, si no esta presente sera cero
        try:
            impte_concepto_p20 = float(hoja.cell_value(fila, 11))
        except ValueError:
            impte_concepto_p20 = 0.0

        # Tomar el importe del concepto PGP, si no esta presente sera cero
        try:
            impte_concepto_pgp = float(hoja.cell_value(fila, 12))
        except ValueError:
            impte_concepto_pgp = 0.0

        # Tomar el importe del concepto PGV, si no esta presente sera cero
        try:
            impte_concepto_pgv = float(hoja.cell_value(fila, 13))
        except ValueError:
            impte_concepto_pgv = 0.0

        # Tomar el importe del concepto D1R, si no esta presente sera cero
        try:
            impte_concepto_d1r = float(hoja.cell_value(fila, 14))
        except ValueError:
            impte_concepto_d1r = 0.0

        # Tomar el importe del concepto D62, si no esta presente sera cero
        try:
            impte_concepto_d62 = float(hoja.cell_value(fila, 15))
        except ValueError:
            impte_concepto_d62 = 0.0

        # Consultar la persona
        persona = Persona.query.filter_by(rfc=rfc).first()

        # Si NO existe, se agrega a la lista de personas_inexistentes y se salta
        if persona is None:
            personas_inexistentes.append(rfc)
            continue

        # Consultar el Centro de Trabajo
        centro_trabajo = CentroTrabajo.query.filter_by(clave=centro_trabajo_clave).first()

        # Si NO existe se agrega a la lista de centros_trabajos_inexistentes y se salta
        if centro_trabajo is None:
            centros_trabajos_inexistentes.append(centro_trabajo_clave)
            continue

        # Consultar la Plaza
        plaza = Plaza.query.filter_by(clave=plaza_clave).first()

        # Si NO existe se agrega a la lista de plazas_inexistentes y se salta
        if plaza is None:
            plazas_inexistentes.append(plaza_clave)
            continue

        # Revisar que en nominas no exista una nomina con la misma persona, quincena y tipo APOYO ANUAL, si existe se omite
        nominas_posibles = (
            Nomina.query.filter_by(persona_id=persona.id)
            .filter_by(quincena_id=quincena.id)
            .filter_by(tipo="PRIMA VACACIONAL")
            .filter_by(estatus="A")
            .all()
        )
        if len(nominas_posibles) > 0:
            nominas_existentes.append(rfc)
            continue

        # Constante tipo
        TIPO = "PRIMA VACACIONAL"

        # Alimentar impte_concepto_p20, con concepto P20
        if impte_concepto_p20 != 0.0:
            if probar is False:
                percepcion_deduccion_p20 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_p20,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_p20,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_p20)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_pgp, con concepto PGP
        if impte_concepto_pgp != 0.0:
            if probar is False:
                percepcion_deduccion_pgp = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pgp,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_pgp,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_pgp)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_pgv, con concepto PGV
        if impte_concepto_pgv != 0.0:
            if probar is False:
                percepcion_deduccion_pgv = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_pgv,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_pgv,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_pgv)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_d1r, con concepto D1R
        if impte_concepto_d1r != 0.0:
            if probar is False:
                percepcion_deduccion_d1r = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d1r,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d1r,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d1r)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar impte_concepto_d62, con concepto D62
        if impte_concepto_d62 != 0.0:
            if probar is False:
                percepcion_deduccion_d62 = PercepcionDeduccion(
                    centro_trabajo=centro_trabajo,
                    concepto=concepto_d62,
                    persona=persona,
                    plaza=plaza,
                    quincena=quincena,
                    importe=impte_concepto_d62,
                    tipo=TIPO,
                )
                sesion.add(percepcion_deduccion_d62)
            click.echo(click.style("p", fg="cyan"), nl=False)

        # Alimentar registro en Nomina
        if probar is False:
            nomina = Nomina(
                centro_trabajo=centro_trabajo,
                persona=persona,
                plaza=plaza,
                quincena=quincena,
                desde=desde,
                desde_clave=desde_clave,
                hasta=hasta,
                hasta_clave=hasta_clave,
                percepcion=percepcion,
                deduccion=deduccion,
                importe=impte,
                tipo=TIPO,
                fecha_pago=fecha_pago,
            )
            sesion.add(nomina)
        click.echo(click.style("a", fg="green"), nl=False)

        # Incrementar contador
        contador += 1

    # Poner avance de linea
    click.echo("")

    # Si contador es cero, mostrar mensaje de error y terminar
    if contador == 0:
        click.echo(click.style("ERROR: No se alimentaron registros en nominas.", fg="red"))
        sys.exit(1)

    # Cerrar la sesion para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Actualizar la quincena para poner en verdadero el campo tiene_primas_vacacionales
    quincena.tiene_primas_vacacionales = True
    quincena.save()

    # Si hubo centros_trabajos_inexistentes, mostrarlos
    if len(centros_trabajos_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(centros_trabajos_inexistentes)} C. de T. que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(centros_trabajos_inexistentes)}", fg="yellow"))

    # Si hubo plazas_inexistentes, mostrarlos
    if len(plazas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(plazas_inexistentes)} Plazas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(plazas_inexistentes)}", fg="yellow"))

    # Si hubo personas_inexistentes, mostrar contador
    if len(personas_inexistentes) > 0:
        click.echo(click.style(f"  Hubo {len(personas_inexistentes)} Personas que no existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(personas_inexistentes)}", fg="yellow"))

    # Si hubo nominas_existentes, mostrarlos
    if len(nominas_existentes) > 0:
        click.echo(click.style(f"  Hubo {len(nominas_existentes)} Primas Vacacionales que ya existen. Se omiten:", fg="yellow"))
        click.echo(click.style(f"  {', '.join(nominas_existentes)}", fg="yellow"))

    # Mensaje termino
    click.echo(
        click.style(f"  Alimentar Primas Vacacionales: {contador} insertadas en la quincena {quincena_clave}.", fg="green")
    )


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--serica-xlsx", default=SERICA_FILENAME_XLSX, help="Archivo XLSX con los datos")
@click.option("--output-txt", default="SERICA.txt", help="Archivo TXT de salida")
def generar_issste(quincena_clave, serica_xlsx, output_txt):
    """Generar archivo XLSX con los datos para el ISSSTE"""

    # Validar si existe el archivo
    ruta = Path(EXPLOTACION_BASE_DIR, quincena_clave, serica_xlsx)
    if not ruta.exists():
        click.echo(f"ERROR: {str(ruta)} no se encontró.")
        sys.exit(1)
    if not ruta.is_file():
        click.echo(f"ERROR: {str(ruta)} no es un archivo.")
        sys.exit(1)

    # Abrir el archivo XLSX con openpyxl para leer sus datos
    workbook = load_workbook(filename=ruta, read_only=True, data_only=True)

    # Elegir la region DETALLE
    detalle = workbook["DETALLE"]

    # Crear archivo TXT con los datos de la fila
    click.echo("Generando ISSSTE: ", nl=False)
    with open(output_txt, "w", encoding="ascii") as f:
        # Bucle por las filas
        for row in detalle.iter_rows(min_row=5, max_col=81, max_row=3000):
            # Juntar todas las celdas de la fila en una lista
            # fila = [str(celda.value).strip() for celda in row]
            fila = [str(celda.value).strip() for celda in row]

            # Si el primer elemento de la fila es vacio, se rompe el ciclo
            if fila[0] == "" or fila[0] is None or fila[0] == "None":
                break

            # Asegurarse de que cada valor sea un string ascii
            for i, valor in enumerate(fila):
                # Si el valor es "0", cambiarlo por "", de lo contrario, convertirlo a ascii
                if valor == "0":
                    fila[i] = ""
                else:
                    fila[i] = valor.encode("ascii", "ignore").decode("ascii")

            # Escribir la fila en el archivo TXT y agregar un salto de linea
            f.write("|".join(fila) + "\r")
            click.echo(click.style(".", fg="cyan"), nl=False)

    # Poner avance de linea
    click.echo("")

    # Mensaje termino
    click.echo(click.style(f"  Generar ISSSTE: {output_txt} generado.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.argument("tipo", type=str)
@click.argument("fecha_pago_str", type=str)
def cambiar_fecha_pago(quincena_clave, tipo, fecha_pago_str):
    """Cambiar las fechas de pago de una quincena y de un tipo de nómina dado"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Validar tipo
    tipo = safe_string(tipo)
    if tipo not in Nomina.TIPOS:
        click.echo(click.style("ERROR: Tipo de nómina inválido.", fg="red"))
        sys.exit(1)

    # Consultar las nóminas de la quincena y del tipo
    nominas = Nomina.query.filter_by(quincena_id=quincena.id).filter_by(tipo=tipo).all()

    # Si no hay nóminas, terminar
    if len(nominas) == 0:
        click.echo(click.style("ERROR: No hay nóminas para cambiar la fecha de pago.", fg="red"))
        sys.exit(1)

    # Validar fecha_pago
    try:
        fecha_pago = datetime.strptime(fecha_pago_str, "%Y-%m-%d").date()
    except ValueError:
        click.echo(click.style("ERROR: Fecha de pago inválida.", fg="red"))
        sys.exit(1)

    # Iniciar sesión con la base de datos para que la alimentación sea rápida
    sesion = database.session

    # Bucle por las nóminas para cambiar la fecha de pago
    contador = 0
    click.echo("Cambiando las fecha de pago: ", nl=False)
    for nomina in nominas:
        if nomina.fecha_pago == fecha_pago:
            continue
        nomina.fecha_pago = fecha_pago
        sesion.add(nomina)
        click.echo(click.style(".", fg="cyan"), nl=False)
        contador += 1

    # Cerrar la sesión para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Poner avance de línea
    click.echo("")

    # Mensaje de termino
    click.echo(click.style(f"Se cambiaron {contador} fechas de pago.", fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--fijar-num-cheque", is_flag=True, help="Definir números de cheque")
@click.option("--modelos-separados-por-comas", help="Por defecto 1,2", default="1,2")
def crear_archivo_xlsx_aguinaldos(quincena_clave, fijar_num_cheque, modelos_separados_por_comas):
    """Crear archivo XLSX con los aguinaldos de una quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="NOMINAS",
        mensajes="Crear archivo XLSX con los aguinaldos de una quincena",
    )
    quincena_producto.save()

    # Ejecutar crear_aguinaldos
    try:
        mensaje_termino = crear_aguinaldos(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            fijar_num_cheque=fijar_num_cheque,
            modelos_separados_por_comas=modelos_separados_por_comas,
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
def crear_archivo_xlsx_dispersiones_pensionados(quincena_clave):
    """Crear archivo XLSX con las dispersiones para pensionados"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="DISPERSIONES PENSIONADOS",
        mensajes="Crear archivo XLSX con las dispersiones para pensionados",
    )
    quincena_producto.save()

    # Ejecutar crear_dispersiones_pensionados
    try:
        mensaje_termino = crear_dispersiones_pensionados(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Mensaje termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--fijar-num-cheque", is_flag=True, help="Definir números de cheque")
def crear_archivo_xlsx_monederos(quincena_clave, fijar_num_cheque):
    """Crear archivo XLSX con los monederos de una quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="MONEDEROS",
        mensajes="Crear archivo XLSX con los monederos de una quincena",
    )
    quincena_producto.save()

    # Ejecutar crear_monederos
    try:
        mensaje_termino = crear_monederos(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            fijar_num_cheque=fijar_num_cheque,
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Mensaje termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--fijar-num-cheque", is_flag=True, help="Definir números de cheque")
def crear_archivo_xlsx_nominas(quincena_clave, fijar_num_chque):
    """Crear archivo XLSX con las nominas de una quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="NOMINAS",
        mensajes="Crear archivo XLSX con las nominas de una quincena",
    )
    quincena_producto.save()

    # Ejecutar crear_nominas
    try:
        mensaje_termino = crear_nominas(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            fijar_num_cheque=fijar_num_chque,
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--fijar-num-cheque", is_flag=True, help="Definir números de cheque")
def crear_archivo_xlsx_pensionados(quincena_clave, fijar_num_cheque):
    """Crear archivo XLSX con los pensionados de una quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="PENSIONADOS",
        mensajes="Crear archivo XLSX con los pensionados de una quincena",
    )
    quincena_producto.save()

    # Ejecutar crear_pensionados
    try:
        mensaje_termino = crear_pensionados(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            fijar_num_cheque=fijar_num_cheque,
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--fijar-num-cheque", is_flag=True, help="Definir números de cheque")
def crear_archivo_xlsx_primas_vacacionales(quincena_clave, fijar_num_cheque):
    """Crear archivo XLSX con las primas vacacionales de una quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="PRIMAS VACACIONALES",
        mensajes="Crear archivo XLSX con las primas vacacionales",
    )
    quincena_producto.save()

    # Ejecutar crear_primas_vacacionales
    try:
        mensaje_termino = crear_primas_vacacionales(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            fijar_num_cheque=fijar_num_cheque,
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
def crear_archivo_xlsx_timbrados_empleados_activos(quincena_clave):
    """Crear archivo XLSX con los timbrados de los empleados activos"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="TIMBRADOS EMPLEADOS ACTIVOS",
        mensajes="Crear archivo XLSX con los timbrados de los empleados activos...",
    )
    quincena_producto.save()

    # Ejecutar crear_timbrados con Personas con modelos 1: "CONFIANZA" y 2: "SINDICALIZADO"
    try:
        mensaje_termino = crear_timbrados(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            modelos=[1, 2],
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
def crear_archivo_xlsx_timbrados_pensionados(quincena_clave):
    """Crear archivo XLSX con los timbrados de los pensionados"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="TIMBRADOS PENSIONADOS",
        mensajes="Crear archivo XLSX con los timbrados de los pensionados",
    )
    quincena_producto.save()

    # Ejecutar crear_timbrados con Personas con modelo 3: "PENSIONADO"
    try:
        mensaje_termino = crear_timbrados(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            modelos=[3],
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
def crear_archivo_xlsx_timbrados_primas_vacacionales(quincena_clave):
    """Crear archivo XLSX con los timbrados de las primas vacacionales"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Validar que la quincena tenga primas vacacionales
    if quincena.tiene_primas_vacacionales is False:
        click.echo(click.style("ERROR: La quincena no tiene primas vacacionales.", fg="red"))
        sys.exit(1)

    # Crear un producto para la quincena
    quincena_producto = QuincenaProducto(
        quincena_id=quincena.id,
        fuente="TIMBRADOS PRIMAS VACACIONALES",
        mensajes="Crear archivo XLSX con los timbrados de las primas vacacionales",
    )
    quincena_producto.save()

    # Ejecutar crear_timbrados con el tipo "PRIMA VACACIONAL"
    try:
        mensaje_termino = crear_timbrados(
            quincena_clave=quincena_clave,
            quincena_producto_id=quincena_producto.id,
            tipo="PRIMA VACACIONAL",
        )
    except Exception as error:
        click.echo(click.style(f"ERROR: {str(error)}", fg="red"))
        sys.exit(1)

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    click.echo(click.style(mensaje_termino, fg="green"))


@click.command()
@click.argument("quincena_clave", type=str)
@click.option("--tipo", help="Tipo de nómina", default="")
def limpiar_num_cheque(quincena_clave: str, tipo: str):
    """Limpiar los números de cheque de una quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        click.echo(click.style("ERROR: Clave de la quincena inválida.", fg="red"))
        sys.exit(1)

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena o ha sido eliminada, causa error
    if quincena is None or quincena.estatus != "A":
        click.echo(click.style("ERROR: No existe o ha sido eliminada la quincena.", fg="red"))
        sys.exit(1)

    # Validar tipo
    if tipo != "":
        tipo = safe_string(tipo)
        if tipo not in Nomina.TIPOS:
            click.echo(click.style("ERROR: Tipo de nómina inválido.", fg="red"))
            sys.exit(1)

    # Consultar las nóminas de la quincena
    nominas = Nomina.query.filter_by(quincena_id=quincena.id)
    if tipo != "":
        nominas = nominas.filter_by(tipo=tipo)
    nominas = nominas.order_by(Nomina.id).all()

    # Si no hay nóminas, terminar
    if len(nominas) == 0:
        click.echo(click.style("ERROR: No hay nóminas para limpiar los números de cheque.", fg="red"))
        sys.exit(1)

    # Iniciar sesión con la base de datos para que la alimentación sea rápida
    sesion = database.session

    # Bucle por las nóminas para cambiar la fecha de pago
    contador = 0
    click.echo("Limpiando los números de cheque: ", nl=False)
    for nomina in nominas:
        if nomina.num_cheque == "":
            continue
        nomina.num_cheque = ""
        sesion.add(nomina)
        click.echo(click.style(".", fg="cyan"), nl=False)
        contador += 1

    # Cerrar la sesión para que se guarden todos los datos en la base de datos
    sesion.commit()
    sesion.close()

    # Poner avance de línea
    click.echo("")

    # Mensaje de termino
    click.echo(click.style(f"Se limpiaron {contador} números de cheque.", fg="green"))


cli.add_command(actualizar_timbrados)
cli.add_command(alimentar)
cli.add_command(alimentar_aguinaldos)
cli.add_command(alimentar_aguinaldos_retroactivos)
cli.add_command(alimentar_apoyos_anuales)
cli.add_command(alimentar_apoyos_madres)
cli.add_command(alimentar_extraordinarios)
cli.add_command(alimentar_pensiones_alimenticias)
cli.add_command(alimentar_primas_vacacionales)
cli.add_command(alimentar_primas_vacacionales_alterno)
cli.add_command(alimentar_salarios_rl)
cli.add_command(generar_issste)
cli.add_command(cambiar_fecha_pago)
cli.add_command(crear_archivo_xlsx_aguinaldos)
cli.add_command(crear_archivo_xlsx_dispersiones_pensionados)
cli.add_command(crear_archivo_xlsx_monederos)
cli.add_command(crear_archivo_xlsx_nominas)
cli.add_command(crear_archivo_xlsx_pensionados)
cli.add_command(crear_archivo_xlsx_primas_vacacionales)
cli.add_command(crear_archivo_xlsx_timbrados_empleados_activos)
cli.add_command(crear_archivo_xlsx_timbrados_pensionados)
cli.add_command(crear_archivo_xlsx_timbrados_primas_vacacionales)
cli.add_command(limpiar_num_cheque)
