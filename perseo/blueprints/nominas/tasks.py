"""
Nominas, tareas en el fondo
"""
import logging
import re
from datetime import datetime
from pathlib import Path

import pytz
from openpyxl import Workbook

from config.settings import get_settings
from lib.exceptions import MyAnyError, MyNotExistsError, MyNotValidParamError
from lib.safe_string import QUINCENA_REGEXP
from lib.storage import GoogleCloudStorage
from lib.tasks import set_task_error, set_task_progress
from perseo.app import create_app
from perseo.blueprints.bancos.models import Banco
from perseo.blueprints.nominas.models import Nomina
from perseo.blueprints.quincenas.models import Quincena
from perseo.blueprints.quincenas_productos.models import QuincenaProducto
from perseo.extensions import database

GCS_BASE_DIRECTORY = "reports/nominas"
LOCAL_BASE_DIRECTORY = "reports/nominas"
TIMEZONE = "America/Mexico_City"

bitacora = logging.getLogger(__name__)
bitacora.setLevel(logging.INFO)
formato = logging.Formatter("%(asctime)s:%(levelname)s:%(message)s")
empunadura = logging.FileHandler("nominas.log")
empunadura.setFormatter(formato)
bitacora.addHandler(empunadura)

app = create_app()
app.app_context().push()
database.app = app


def consultar_validar_quincena(quincena_clave: str) -> Quincena:
    """Consultar y validar la quincena"""

    # Validar quincena_clave
    if re.match(QUINCENA_REGEXP, quincena_clave) is None:
        raise MyNotValidParamError("Clave de la quincena inválida.")

    # Consultar quincena
    quincena = Quincena.query.filter_by(clave=quincena_clave).first()

    # Si no existe la quincena, provocar error y terminar
    if quincena is None:
        return MyNotExistsError(f"No existe la quincena {quincena_clave}")

    # Si la quincena no esta ABIERTA, provocar error y terminar
    if quincena.estado != "ABIERTA":
        return MyNotValidParamError(f"La quincena {quincena_clave} no esta ABIERTA")

    # Si la quincena esta eliminada, provocar error y terminar
    if quincena.estatus != "A":
        return MyNotValidParamError(f"La quincena {quincena_clave} esta eliminada")

    # Entregar la quincena
    return quincena


def crear_nominas(quincena_clave: str, quincena_producto_id: int) -> str:
    """Crear archivo XLSX con las nominas de una quincena"""

    # Consultar y validar quincena
    quincena = consultar_validar_quincena(quincena_clave)  # Puede provocar una excepcion

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Consultar las nominas de la quincena, solo tipo SALARIO
    nominas = Nomina.query.filter_by(quincena_id=quincena.id).filter_by(tipo="SALARIO").filter_by(estatus="A").all()

    # Si no hay nominas, provocar error y salir
    if len(nominas) == 0:
        raise MyNotExistsError(f"No hay nominas de tipo SALARIO en la quincena {quincena_clave}")

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "QUINCENA",
            "CENTRO DE TRABAJO",
            "RFC",
            "NOMBRE COMPLETO",
            "NUMERO DE EMPLEADO",
            "MODELO",
            "PLAZA",
            "NOMBRE DEL BANCO",
            "BANCO ADMINISTRADOR",
            "NUMERO DE CUENTA",
            "MONTO A DEPOSITAR",
            "NO DE CHEQUE",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Si el modelo de la persona es 3, se omite
        if nomina.persona.modelo == 3:
            continue

        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave != "9" and cuenta.estatus == "A":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar el banco de la cuenta de la persona
        su_banco = su_cuenta.banco

        # Incrementar la consecutivo del banco
        su_banco.consecutivo_generado += 1

        # Elaborar el numero de cheque, juntando la clave del banco y la consecutivo, siempre de 9 digitos
        num_cheque = f"{su_cuenta.banco.clave.zfill(2)}{su_banco.consecutivo_generado:07}"

        # Agregar la fila
        hoja.append(
            [
                nomina.quincena.clave,
                nomina.centro_trabajo.clave,
                nomina.persona.rfc,
                nomina.persona.nombre_completo,
                nomina.persona.num_empleado,
                nomina.persona.modelo,
                nomina.plaza.clave,
                su_banco.nombre,
                su_banco.clave,
                su_cuenta.num_cuenta,
                nomina.importe,
                num_cheque,
            ]
        )

        # Incrementar contador
        contador += 1

    # Actualizar los consecutivos de cada banco
    sesion.commit()

    # Determinar la fecha y tiempo actual en la zona horaria de Mexico
    ahora = datetime.now(tz=pytz.timezone(TIMEZONE))

    # Determinar el nombre y ruta del archivo XLSX
    nombre_archivo_xlsx = f"nominas_{quincena_clave}_{ahora.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    descripcion_archivo_xlsx = f"Nominas {quincena_clave} {ahora.strftime('%Y-%m-%d %H%M%S')}"
    archivo_xlsx = Path(LOCAL_BASE_DIRECTORY, nombre_archivo_xlsx)

    # Si no existe la carpeta LOCAL_BASE_DIRECTORY, crearla
    Path(LOCAL_BASE_DIRECTORY).mkdir(parents=True, exist_ok=True)

    # Guardar el archivo XLSX
    libro.save(archivo_xlsx)

    # Si esta configurado settings.CLOUD_STORAGE_DEPOSITO, entonces subir el archivo XLSX a Google Cloud Storage
    gcs_public_path = ""
    settings = get_settings()
    if settings.CLOUD_STORAGE_DEPOSITO != "":
        with open(archivo_xlsx, "rb") as archivo:
            try:
                bitacora.info("GCS: Bucket %s", settings.CLOUD_STORAGE_DEPOSITO)
                gcstorage = GoogleCloudStorage(
                    base_directory=GCS_BASE_DIRECTORY,
                    upload_date=ahora.date(),
                    allowed_extensions=["xlsx"],
                    month_in_word=False,
                    bucket_name=settings.CLOUD_STORAGE_DEPOSITO,
                )
                gcs_nombre_archivo_xlsx = gcstorage.set_filename(
                    description=descripcion_archivo_xlsx,
                    extension="xlsx",
                    start_with_date=False,
                )
                bitacora.info("GCS: Subiendo %s", gcs_nombre_archivo_xlsx)
                gcs_public_path = gcstorage.upload(archivo.read())
                bitacora.info("GCS: Depositado %s", gcs_public_path)
            except MyAnyError as error:
                raise error

    # Si hubo personas sin cuentas, entonces juntarlas para mensajes
    mensajes = []
    if len(personas_sin_cuentas) > 0:
        mensajes.append(f"AVISO: Hubo {len(personas_sin_cuentas)} personas sin cuentas:")
        mensajes += [f"- {p.rfc} {p.nombre_completo}" for p in personas_sin_cuentas]
    if len(mensajes) > 0:
        for m in mensajes:
            bitacora.warning(m)

    # Si quincena_producto_id es cero, agregar un registro para conservar las rutas y mensajes
    if quincena_producto_id == 0:
        quincena_producto = QuincenaProducto(
            quincena=quincena,
            archivo=nombre_archivo_xlsx,
            es_satisfactorio=(len(mensajes) == 0),
            fuente="NOMINAS",
            mensajes="\n".join(mensajes),
            url=gcs_public_path,
        )
    else:
        # Si quincena_producto_id es diferente de cero, actualizar el registro
        quincena_producto = QuincenaProducto.query.get(quincena_producto_id)
        quincena_producto.archivo = nombre_archivo_xlsx
        quincena_producto.es_satisfactorio = len(mensajes) == 0
        quincena_producto.fuente = "NOMINAS"
        quincena_producto.mensajes = "\n".join(mensajes)
        quincena_producto.url = gcs_public_path
    quincena_producto.save()

    # Entregar mensaje de termino
    return f"Generar nominas: {contador} filas en {nombre_archivo_xlsx}"


def generar_nominas(quincena_clave: str, quincena_producto_id: int) -> str:
    """Tarea en el fondo para crear un archivo XLSX con las nominas de una quincena"""

    # Iniciar la tarea en el fondo
    set_task_progress(0, f"Generar archivo XLSX con las nominas de {quincena_clave}...")

    # Ejecutar el creador
    try:
        mensaje_termino = crear_nominas(quincena_clave, quincena_producto_id)
    except MyAnyError as error:
        mensaje_error = str(error)
        set_task_error(mensaje_error)
        bitacora.error(mensaje_error)
        return mensaje_error

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    set_task_progress(100, mensaje_termino)
    bitacora.info(mensaje_termino)
    return mensaje_termino


def crear_monederos(quincena_clave: str, quincena_producto_id: int) -> str:
    """Crear archivo XLSX con los monederos de una quincena"""

    # Consultar y validar quincena
    quincena = consultar_validar_quincena(quincena_clave)  # Puede provocar una excepcion

    # Iniciar sesion con la base de datos para que la alimentacion sea rapida
    sesion = database.session

    # Cargar solo el banco con la clave 9 que es PREVIVALE
    banco = Banco.query.filter_by(clave="9").first()

    # Si no existe el banco, provocar error y salir
    if banco is None:
        raise MyNotExistsError("No existe el banco con clave 9")

    # Igualar el consecutivo_generado al consecutivo
    banco.consecutivo_generado = banco.consecutivo

    # Consultar las nominas de la quincena solo tipo DESPENSA
    nominas = Nomina.query.filter_by(quincena_id=quincena.id).filter_by(tipo="DESPENSA").filter_by(estatus="A").all()

    # Si no hay nominas, provocar error y salir
    if len(nominas) == 0:
        raise MyNotExistsError(f"No hay nominas de tipo SALARIO en la quincena {quincena_clave}")

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "CT_CLASIF",
            "RFC",
            "TOT NET CHEQUE",
            "NUM CHEQUE",
            "NUM TARJETA",
            "QUINCENA",
            "MODELO",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave == "9" and cuenta.estatus == "A":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se agrega a la lista de personas_sin_cuentas y se salta
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Incrementar el consecutivo_generado del banco
        banco.consecutivo_generado += 1

        # Elaborar el numero de cheque, juntando la clave del banco y el consecutivo_generado, siempre de 9 digitos
        num_cheque = f"{su_cuenta.banco.clave.zfill(2)}{banco.consecutivo_generado:07}"

        # Agregar la fila
        hoja.append(
            [
                "J",
                nomina.persona.rfc,
                nomina.importe,
                num_cheque,
                su_cuenta.num_cuenta,
                nomina.quincena.clave,
                nomina.persona.modelo,
            ]
        )

        # Incrementar contador
        contador += 1

    # Actualizar los consecutivo_generado de cada banco
    sesion.commit()

    # Determinar la fecha y tiempo actual en la zona horaria de Mexico
    ahora = datetime.now(tz=pytz.timezone(TIMEZONE))

    # Determinar el nombre y ruta del archivo XLSX
    nombre_archivo_xlsx = f"monederos_{quincena_clave}_{ahora.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    descripcion_archivo_xlsx = f"Monederos {quincena_clave} {ahora.strftime('%Y-%m-%d %H%M%S')}"
    archivo_xlsx = Path(LOCAL_BASE_DIRECTORY, nombre_archivo_xlsx)

    # Si no existe la carpeta LOCAL_BASE_DIRECTORY, crearla
    Path(LOCAL_BASE_DIRECTORY).mkdir(parents=True, exist_ok=True)

    # Guardar el archivo XLSX
    libro.save(archivo_xlsx)

    # Si esta configurado settings.CLOUD_STORAGE_DEPOSITO, entonces subir el archivo XLSX a Google Cloud Storage
    settings = get_settings()
    if settings.CLOUD_STORAGE_DEPOSITO != "":
        with open(archivo_xlsx, "rb") as archivo:
            try:
                bitacora.info("GCS: Bucket %s", settings.CLOUD_STORAGE_DEPOSITO)
                gcstorage = GoogleCloudStorage(
                    base_directory=GCS_BASE_DIRECTORY,
                    upload_date=ahora.date(),
                    allowed_extensions=["xlsx"],
                    month_in_word=False,
                    bucket_name=settings.CLOUD_STORAGE_DEPOSITO,
                )
                gcs_nombre_archivo_xlsx = gcstorage.set_filename(
                    description=descripcion_archivo_xlsx,
                    extension="xlsx",
                    start_with_date=False,
                )
                bitacora.info("GCS: Subiendo %s", gcs_nombre_archivo_xlsx)
                gcs_public_path = gcstorage.upload(archivo.read())
                bitacora.info("GCS: Depositado %s", gcs_public_path)
            except MyAnyError as error:
                raise error

    # Si hubo personas sin cuentas, entonces juntarlas para mensajes
    mensajes = []
    if len(personas_sin_cuentas) > 0:
        mensajes.append(f"AVISO: Hubo {len(personas_sin_cuentas)} personas sin cuentas:")
        mensajes += [f"- {p.rfc} {p.nombre_completo}" for p in personas_sin_cuentas]
    if len(mensajes) > 0:
        for m in mensajes:
            bitacora.warning(m)

    # Si quincena_producto_id es cero, agregar un registro para conservar las rutas y mensajes
    if quincena_producto_id == 0:
        quincena_producto = QuincenaProducto(
            quincena=quincena,
            archivo=nombre_archivo_xlsx,
            es_satisfactorio=(len(mensajes) == 0),
            fuente="MONEDEROS",
            mensajes="\n".join(mensajes),
            url=gcs_public_path,
        )
    else:
        # Si quincena_producto_id es diferente de cero, actualizar el registro
        quincena_producto = QuincenaProducto.query.get(quincena_producto_id)
        quincena_producto.archivo = nombre_archivo_xlsx
        quincena_producto.es_satisfactorio = len(mensajes) == 0
        quincena_producto.fuente = "MONEDEROS"
        quincena_producto.mensajes = "\n".join(mensajes)
        quincena_producto.url = gcs_public_path
    quincena_producto.save()

    # Entregar mensaje de termino
    return f"Generar monederos: {contador} filas en {nombre_archivo_xlsx}"


def generar_monederos(quincena_clave: str, quincena_producto_id: int) -> str:
    """Tarea en el fondo para crear un archivo XLSX con los monederos de una quincena"""

    # Iniciar la tarea en el fondo
    set_task_progress(0, f"Generar archivo XLSX con los monederos de {quincena_clave}...")

    # Ejecutar el creador
    try:
        mensaje_termino = crear_monederos(quincena_clave, quincena_producto_id)
    except MyAnyError as error:
        mensaje_error = str(error)
        set_task_error(mensaje_error)
        bitacora.error(mensaje_error)
        return mensaje_error

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    set_task_progress(100, mensaje_termino)
    bitacora.info(mensaje_termino)
    return mensaje_termino


def crear_pensionados(quincena_clave: str, quincena_producto_id: int) -> str:
    """Crear archivo XLSX con los pensionados de una quincena"""

    # Consultar y validar quincena
    quincena = consultar_validar_quincena(quincena_clave)  # Puede provocar una excepcion

    # Consultar las nominas de la quincena, solo tipo SALARIO
    nominas = Nomina.query.filter_by(quincena_id=quincena.id).filter_by(tipo="SALARIO").filter_by(estatus="A").all()

    # Si no hay nominas, provocar error y salir
    if len(nominas) == 0:
        raise MyNotExistsError(f"No hay nominas de tipo SALARIO en la quincena {quincena_clave}")

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "QUINCENA",
            "CENTRO DE TRABAJO",
            "RFC",
            "NOMBRE COMPLETO",
            "NUMERO DE EMPLEADO",
            "MODELO",
            "PLAZA",
            "NOMBRE DEL BANCO",
            "BANCO ADMINISTRADOR",
            "NUMERO DE CUENTA",
            "MONTO A DEPOSITAR",
            "NO DE CHEQUE",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Si el modelo de la persona NO es 3, se omite
        if nomina.persona.modelo != 3:
            continue

        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se le crea una cuenta con el banco 10
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave != "9" and cuenta.estatus == "A":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se le crea una cuenta con el banco 10
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar el banco de la cuenta de la persona
        su_banco = su_cuenta.banco

        # Incrementar la consecutivo del banco
        su_banco.consecutivo_generado += 1

        # Elaborar el numero de cheque, juntando la clave del banco y la consecutivo, siempre de 9 digitos
        num_cheque = f"{su_cuenta.banco.clave.zfill(2)}{su_banco.consecutivo_generado:07}"

        # Agregar la fila
        hoja.append(
            [
                nomina.quincena.clave,
                nomina.centro_trabajo.clave,
                nomina.persona.rfc,
                nomina.persona.nombre_completo,
                nomina.persona.num_empleado,
                nomina.persona.modelo,
                nomina.plaza.clave,
                su_banco.nombre,
                su_banco.clave,
                su_cuenta.num_cuenta,
                nomina.importe,
                num_cheque,
            ]
        )

        # Incrementar contador
        contador += 1

    # Determinar la fecha y tiempo actual en la zona horaria de Mexico
    ahora = datetime.now(tz=pytz.timezone(TIMEZONE))

    # Determinar el nombre y ruta del archivo XLSX
    nombre_archivo_xlsx = f"pensionados_{quincena_clave}_{ahora.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    descripcion_archivo_xlsx = f"Pensionados {quincena_clave} {ahora.strftime('%Y-%m-%d %H%M%S')}"
    archivo_xlsx = Path(LOCAL_BASE_DIRECTORY, nombre_archivo_xlsx)

    # Si no existe la carpeta LOCAL_BASE_DIRECTORY, crearla
    Path(LOCAL_BASE_DIRECTORY).mkdir(parents=True, exist_ok=True)

    # Guardar el archivo XLSX
    libro.save(archivo_xlsx)

    # Si esta configurado settings.CLOUD_STORAGE_DEPOSITO, entonces subir el archivo XLSX a Google Cloud Storage
    settings = get_settings()
    if settings.CLOUD_STORAGE_DEPOSITO != "":
        with open(archivo_xlsx, "rb") as archivo:
            try:
                bitacora.info("GCS: Bucket %s", settings.CLOUD_STORAGE_DEPOSITO)
                gcstorage = GoogleCloudStorage(
                    base_directory=GCS_BASE_DIRECTORY,
                    upload_date=ahora.date(),
                    allowed_extensions=["xlsx"],
                    month_in_word=False,
                    bucket_name=settings.CLOUD_STORAGE_DEPOSITO,
                )
                gcs_nombre_archivo_xlsx = gcstorage.set_filename(
                    description=descripcion_archivo_xlsx,
                    extension="xlsx",
                    start_with_date=False,
                )
                bitacora.info("GCS: Subiendo %s", gcs_nombre_archivo_xlsx)
                gcs_public_path = gcstorage.upload(archivo.read())
                bitacora.info("GCS: Depositado %s", gcs_public_path)
            except MyAnyError as error:
                raise error

    # Si hubo personas sin cuentas, entonces juntarlas para mensajes
    mensajes = []
    if len(personas_sin_cuentas) > 0:
        mensajes.append(f"AVISO: Hubo {len(personas_sin_cuentas)} personas sin cuentas:")
        mensajes += [f"- {p.rfc} {p.nombre_completo}" for p in personas_sin_cuentas]
    if len(mensajes) > 0:
        for m in mensajes:
            bitacora.warning(m)

    # Si quincena_producto_id es cero, agregar un registro para conservar las rutas y mensajes
    if quincena_producto_id == 0:
        quincena_producto = QuincenaProducto(
            quincena=quincena,
            archivo=nombre_archivo_xlsx,
            es_satisfactorio=(len(mensajes) == 0),
            fuente="PENSIONADOS",
            mensajes="\n".join(mensajes),
            url=gcs_public_path,
        )
    else:
        # Si quincena_producto_id es diferente de cero, actualizar el registro
        quincena_producto = QuincenaProducto.query.get(quincena_producto_id)
        quincena_producto.archivo = nombre_archivo_xlsx
        quincena_producto.es_satisfactorio = len(mensajes) == 0
        quincena_producto.fuente = "PENSIONADOS"
        quincena_producto.mensajes = "\n".join(mensajes)
        quincena_producto.url = gcs_public_path
    quincena_producto.save()

    # Entregar mensaje de termino
    return f"Generar pensionados: {contador} filas en {nombre_archivo_xlsx}"


def generar_pensionados(quincena_clave: str, quincena_producto_id: int) -> str:
    """Tarea en el fondo para crear un archivo XLSX con los pensionados de una quincena"""

    # Iniciar la tarea en el fondo
    set_task_progress(0, f"Generar archivo XLSX con los pensionados de {quincena_clave}...")

    # Ejecutar el creador
    try:
        mensaje_termino = crear_pensionados(quincena_clave, quincena_producto_id)
    except MyAnyError as error:
        mensaje_error = str(error)
        set_task_error(mensaje_error)
        bitacora.error(mensaje_error)
        return mensaje_error

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    set_task_progress(100, mensaje_termino)
    bitacora.info(mensaje_termino)
    return mensaje_termino


def crear_dispersiones_pensionados(quincena_clave: str, quincena_producto_id: int) -> str:
    """Crear archivo XLSX con las dispersiones pensionados de una quincena"""

    # Consultar y validar quincena
    quincena = consultar_validar_quincena(quincena_clave)  # Puede provocar una excepcion

    # Consultar las nominas de la quincena, solo tipo SALARIO
    nominas = Nomina.query.filter_by(quincena_id=quincena.id).filter_by(tipo="SALARIO").filter_by(estatus="A").all()

    # Si no hay nominas, provocar error y salir
    if len(nominas) == 0:
        raise MyNotExistsError(f"No hay nominas de tipo SALARIO en la quincena {quincena_clave}")

    # Iniciar el archivo XLSX
    libro = Workbook()

    # Tomar la hoja del libro XLSX
    hoja = libro.active

    # Agregar la fila con las cabeceras de las columnas
    hoja.append(
        [
            "CONSECUTIVO",
            "FORMA DE PAGO",
            "TIPO DE CUENTA",
            "BANCO RECEPTOR",
            "CUENTA ABONO",
            "IMPORTE PAGO",
            "CLAVE BENEFICIARIO",
            "RFC",
            "NOMBRE",
            "REFERENCIA PAGO",
            "CONCEPTO PAGO",
        ]
    )

    # Bucle para crear cada fila del archivo XLSX
    contador = 0
    personas_sin_cuentas = []
    for nomina in nominas:
        # Si el modelo de la persona NO es 3, se omite
        if nomina.persona.modelo != 3:
            continue

        # Tomar las cuentas de la persona
        cuentas = nomina.persona.cuentas

        # Si no tiene cuentas, entonces se le crea una cuenta con el banco 10
        if len(cuentas) == 0:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Tomar la cuenta de la persona que no tenga la clave 9, porque esa clave es la de DESPENSA
        su_cuenta = None
        for cuenta in cuentas:
            if cuenta.banco.clave != "9" and cuenta.estatus == "A":
                su_cuenta = cuenta
                break

        # Si no tiene cuenta bancaria, entonces se le crea una cuenta con el banco 10
        if su_cuenta is None:
            personas_sin_cuentas.append(nomina.persona)
            continue

        # Definir referencia_pago, se forma con los dos ultimos caracteres y los caracteres tercero y cuarto de la quincena
        referencia_pago = f"{quincena_clave[-2:]}{quincena_clave[2:4]}"

        # Definir concepto_pago, se forma con el texto "QUINCENA {dos digitos} PENSIONADOS"
        concepto_pago = f"QUINCENA {quincena_clave[-2:]} PENSIONADOS"

        # Agregar la fila
        hoja.append(
            [
                contador + 1,
                "04",
                "9",
                su_cuenta.banco.clave_dispersion_pensionados,
                su_cuenta.num_cuenta,
                nomina.importe,
                contador + 1,
                nomina.persona.rfc,
                nomina.persona.nombre_completo,
                referencia_pago,
                concepto_pago,
            ]
        )

        # Incrementar contador
        contador += 1

    # Determinar la fecha y tiempo actual en la zona horaria de Mexico
    ahora = datetime.now(tz=pytz.timezone(TIMEZONE))

    # Determinar el nombre y ruta del archivo XLSX
    nombre_archivo_xlsx = f"dispersiones_pensionados_{quincena_clave}_{ahora.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    descripcion_archivo_xlsx = f"Dispersiones Pensionados {quincena_clave} {ahora.strftime('%Y-%m-%d %H%M%S')}"
    archivo_xlsx = Path(LOCAL_BASE_DIRECTORY, nombre_archivo_xlsx)

    # Si no existe la carpeta LOCAL_BASE_DIRECTORY, crearla
    Path(LOCAL_BASE_DIRECTORY).mkdir(parents=True, exist_ok=True)

    # Guardar el archivo XLSX
    libro.save(archivo_xlsx)

    # Si esta configurado settings.CLOUD_STORAGE_DEPOSITO, entonces subir el archivo XLSX a Google Cloud Storage
    settings = get_settings()
    if settings.CLOUD_STORAGE_DEPOSITO != "":
        with open(archivo_xlsx, "rb") as archivo:
            try:
                bitacora.info("GCS: Bucket %s", settings.CLOUD_STORAGE_DEPOSITO)
                gcstorage = GoogleCloudStorage(
                    base_directory=GCS_BASE_DIRECTORY,
                    upload_date=ahora.date(),
                    allowed_extensions=["xlsx"],
                    month_in_word=False,
                    bucket_name=settings.CLOUD_STORAGE_DEPOSITO,
                )
                gcs_nombre_archivo_xlsx = gcstorage.set_filename(
                    description=descripcion_archivo_xlsx,
                    extension="xlsx",
                    start_with_date=False,
                )
                bitacora.info("GCS: Subiendo %s", gcs_nombre_archivo_xlsx)
                gcs_public_path = gcstorage.upload(archivo.read())
                bitacora.info("GCS: Depositado %s", gcs_public_path)
            except MyAnyError as error:
                raise error

    # Si hubo personas sin cuentas, entonces juntarlas para mensajes
    mensajes = []
    if len(personas_sin_cuentas) > 0:
        mensajes.append(f"AVISO: Hubo {len(personas_sin_cuentas)} personas sin cuentas:")
        mensajes += [f"- {p.rfc} {p.nombre_completo}" for p in personas_sin_cuentas]
    if len(mensajes) > 0:
        for m in mensajes:
            bitacora.warning(m)

    # Si quincena_producto_id es cero, agregar un registro para conservar las rutas y mensajes
    if quincena_producto_id == 0:
        quincena_producto = QuincenaProducto(
            quincena=quincena,
            archivo=nombre_archivo_xlsx,
            es_satisfactorio=(len(mensajes) == 0),
            fuente="DISPERSIONES PENSIONADOS",
            mensajes="\n".join(mensajes),
            url=gcs_public_path,
        )
    else:
        # Si quincena_producto_id es diferente de cero, actualizar el registro
        quincena_producto = QuincenaProducto.query.get(quincena_producto_id)
        quincena_producto.archivo = nombre_archivo_xlsx
        quincena_producto.es_satisfactorio = len(mensajes) == 0
        quincena_producto.fuente = "DISPERSIONES PENSIONADOS"
        quincena_producto.mensajes = "\n".join(mensajes)
        quincena_producto.url = gcs_public_path
    quincena_producto.save()

    # Entregar mensaje de termino
    return f"Generar dispersiones pensionados: {contador} filas en {nombre_archivo_xlsx}"


def generar_dispersiones_pensionados(quincena_clave: str, quincena_producto_id: int) -> str:
    """Tarea en el fondo para crear un archivo XLSX con las dispersiones pensionados de una quincena"""

    # Iniciar la tarea en el fondo
    set_task_progress(0, f"Generar archivo XLSX con las dispersiones pensionados de {quincena_clave}...")

    # Ejecutar el creador
    try:
        mensaje_termino = crear_dispersiones_pensionados(quincena_clave, quincena_producto_id)
    except MyAnyError as error:
        mensaje_error = str(error)
        set_task_error(mensaje_error)
        bitacora.error(mensaje_error)
        return mensaje_error

    # Terminar la tarea en el fondo y entregar el mensaje de termino
    set_task_progress(100, mensaje_termino)
    bitacora.info(mensaje_termino)
    return mensaje_termino


def generar_todos(quincena_clave: str) -> str:
    """Generar todos los archivos XLSX de una quincena"""

    # Iniciar la tarea en el fondo
    set_task_progress(0, f"Generar todos los archivos XLSX de {quincena_clave}...")

    # Ejecutar cada uno de los generadores
    try:
        mensaje_1 = crear_nominas(quincena_clave, 0)
        set_task_progress(25, mensaje_1)
        mensaje_2 = crear_monederos(quincena_clave, 0)
        set_task_progress(50, mensaje_2)
        mensaje_3 = crear_pensionados(quincena_clave, 0)
        set_task_progress(75, mensaje_3)
        mensaje_4 = crear_dispersiones_pensionados(quincena_clave, 0)
        set_task_progress(100, mensaje_4)
    except MyAnyError as error:
        mensaje_error = str(error)
        set_task_error(mensaje_error)
        bitacora.error(mensaje_error)

    # Entregar mensaje de termino
    mensaje_termino = f"{mensaje_1}. \n{mensaje_2}. \n{mensaje_3}. \n{mensaje_4}"
    set_task_progress(100, mensaje_termino)
    bitacora.info(mensaje_termino)
    return mensaje_termino
